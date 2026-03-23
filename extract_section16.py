"""
extract_section16.py
Scans .docx files in 'input/' (recursive), extracts Section 16 tables,
writes formatted results to 'output/Section16_Extract.xlsx'.

Columns: Folder Name | Word Doc Name | Section Name | Table Content (dynamic)

Requirements: pip install python-docx openpyxl
"""

import sys
from pathlib import Path
from dataclasses import dataclass, field

try:
    from docx import Document
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
except ImportError:
    sys.exit("Missing dependency: pip install python-docx")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

INPUT_DIR  = Path(__file__).parent / "input"
OUTPUT_DIR = Path(__file__).parent / "output"

# ── Styles (module-level constants) ──────────────────────────────────────────
_THIN   = Side(style="thin",   color="BFBFBF")
_MEDIUM = Side(style="medium", color="4472C4")
_BRD    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BRD_META    = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)
_BRD_DIVIDER = Border(left=_THIN,   right=_THIN,  bottom=_THIN,
                      top=Side(style="medium", color="1F4E79"))

def _style(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border

_F_SHEET_HDR = Font(name="Arial", bold=True, color="FFFFFF",  size=11)
_F_TBL_HDR   = Font(name="Arial", bold=True, color="FFFFFF",  size=10)
_F_DATA      = Font(name="Arial", size=10)
_F_META      = Font(name="Arial", bold=True, color="1F4E79",  size=10)

_FILL_SHEET_HDR = PatternFill("solid", fgColor="1F4E79")
_FILL_TBL_HDR   = PatternFill("solid", fgColor="2E75B6")
_FILL_ALT       = PatternFill("solid", fgColor="D9E8F5")
_FILL_META      = PatternFill("solid", fgColor="F2F2F2")
_FILL_NONE      = PatternFill()

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_TOP    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
_ALIGN_MID    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Data model ────────────────────────────────────────────────────────────────
@dataclass
class DocRecord:
    folder:      str
    filename:    str
    section:     str
    table_rows:  list = field(default_factory=list)  # list[list[str] | None]

# ── Extraction ────────────────────────────────────────────────────────────────
def _pstyle(elem) -> str:
    ps = elem.find(".//" + qn("w:pStyle"))
    return ps.get(qn("w:val")) if ps is not None else ""

def _text(elem) -> str:
    return "".join(t.text or "" for t in elem.findall(".//" + qn("w:t"))).strip()

def _parse_table(tbl_elem, doc) -> list[list[str]]:
    rows = []
    for row in DocxTable(tbl_elem, doc).rows:
        seen, cells = set(), []
        for cell in row.cells:
            if id(cell) not in seen:
                seen.add(id(cell))
                cells.append(cell.text.replace("\n", " ").strip())
        rows.append(cells)
    return rows

def extract_section16(doc_path: Path) -> tuple[str, list]:
    try:
        doc = Document(str(doc_path))
    except Exception as exc:
        return f"ERROR: {exc}", []

    section, rows, in_s16, h1_count, first = "", [], False, 0, True
    for elem in doc.element.body:
        tag = elem.tag.split("}")[-1]
        if tag == "p" and _pstyle(elem) == "Heading1":
            h1_count += 1
            if h1_count == 16:
                in_s16, section = True, "16\t" + _text(elem)
            elif h1_count == 17 and in_s16:
                break
        elif tag == "tbl" and in_s16:
            tbl_rows = _parse_table(elem, doc)
            if tbl_rows:
                if not first:
                    rows.append(None)
                rows.extend(tbl_rows)
                first = False

    return (section or "Section 16 Not Found"), rows

# ── Excel writer ──────────────────────────────────────────────────────────────
META_COLS = 3   # Folder | Doc Name | Section Name

def _write_meta_cells(ws, excel_row, values, border=_BRD_META):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=excel_row, column=col, value=val)
        _style(c, font=_F_META, fill=_FILL_META, align=_ALIGN_TOP, border=border)

def _fill_meta_cols(ws, excel_row, fill, border):
    for col in range(1, META_COLS + 1):
        c = ws.cell(row=excel_row, column=col)
        c.fill, c.border = fill, border

def _merge_meta(ws, start_row, end_row, values):
    for col, val in enumerate(values, start=1):
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row,     end_column=col)
        c = ws.cell(row=start_row, column=col, value=val)
        _style(c, font=_F_META, fill=_FILL_META, align=_ALIGN_TOP, border=_BRD_META)

def write_excel(records: list[DocRecord], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Section 16 Extract"

    # Header row
    for col, label in enumerate(["Folder Name", "Word Doc Name", "Section Name", "Table Content"], start=1):
        c = ws.cell(row=1, column=col, value=label)
        _style(c, font=_F_SHEET_HDR, fill=_FILL_SHEET_HDR, align=_ALIGN_CENTER, border=_BRD)
    ws.row_dimensions[1].height = 28

    excel_row = 2
    for rec in records:
        meta_vals = [rec.folder, rec.filename, rec.section]

        if not rec.table_rows:
            _write_meta_cells(ws, excel_row, meta_vals)
            c = ws.cell(row=excel_row, column=META_COLS + 1, value="(No tables found in Section 16)")
            _style(c, font=_F_DATA, align=_ALIGN_TOP, border=_BRD)
            excel_row += 1
            continue

        doc_start   = excel_row
        alt         = 0
        is_hdr      = True
        need_divider = False

        for trow in rec.table_rows:
            if trow is None:
                is_hdr, alt, need_divider = True, 0, True
                continue

            brd_data = _BRD_DIVIDER if (is_hdr and need_divider) else _BRD
            fill_row = _FILL_TBL_HDR if is_hdr else (_FILL_ALT if alt % 2 else _FILL_NONE)
            font_row = _F_TBL_HDR   if is_hdr else _F_DATA
            align_row = _ALIGN_MID  if is_hdr else _ALIGN_TOP

            _fill_meta_cols(ws, excel_row, fill_row, brd_data)
            for offset, text in enumerate(trow):
                c = ws.cell(row=excel_row, column=META_COLS + 1 + offset, value=text)
                _style(c, font=font_row, fill=fill_row, align=align_row, border=brd_data)

            if not is_hdr:
                alt += 1
            ws.row_dimensions[excel_row].height = 18
            is_hdr, need_divider = False, False
            excel_row += 1

        _merge_meta(ws, doc_start, excel_row - 1, meta_vals)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 38
    for col in range(META_COLS + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 32

    ws.freeze_panes = "D2"
    wb.save(str(output_path))

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not INPUT_DIR.exists():
        sys.exit(f"Input folder not found: {INPUT_DIR}")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    docx_files = sorted(p for p in INPUT_DIR.rglob("*.docx") if not p.name.startswith("~$"))
    if not docx_files:
        sys.exit(f"No .docx files found under: {INPUT_DIR}")

    print(f"Found {len(docx_files)} Word file(s). Extracting Section 16 tables …\n")

    records = []
    for doc_path in docx_files:
        rel = doc_path.relative_to(INPUT_DIR)
        folder = "" if str(rel.parent) == "." else str(rel.parent)
        section, table_rows = extract_section16(doc_path)
        n = sum(1 for r in table_rows if r is not None)
        print(f"  {'✓' if table_rows else '–'}  {rel}  →  {section[:60]}  ({n} rows)")
        records.append(DocRecord(folder=folder, filename=doc_path.name,
                                 section=section, table_rows=table_rows))

    output_file = OUTPUT_DIR / "Section16_Extract.xlsx"
    write_excel(records, output_file)
    print(f"\nDone! Output saved to: {output_file}")

if __name__ == "__main__":
    main()