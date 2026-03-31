"""
Linked Object Scanner — scans .docx AND .doc files in Input/ for embedded/linked OLE objects.
Objects with a URL are recorded; objects without are extracted to Output/<DocName>/.
Run: python scan_linked_objects.py

Report columns (8 total):
  A  Folder                      ← immediate parent folder of the file
  B  Word Document
  C  Word Section                ← section number + name (e.g. "17. APPENDIX")
  D  Row Label
  E  Row Label Column Header     ← header of the column that contains the row label
  F  Object Type
  G  URL                         ← hyperlink URL or relative path to extracted file
  H  URL Column Header           ← header of the column that contains the OLE object
"""

import os, io, re, struct, shutil, zipfile, urllib.parse
from collections import defaultdict
from datetime import datetime
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_HERE         = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER  = os.path.join(_HERE, "Input")
OUTPUT_FOLDER = os.path.join(_HERE, "Output")
REPORT_NAME   = "embedded_objects_report.xlsx"

NS_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_W   = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
NS_V   = 'urn:schemas-microsoft-com:vml'


# ── Utilities ─────────────────────────────────────────────────────────────────

def ole_type(prog, ext):
    if 'Excel'   in prog or ext in ('.xlsx','.xls','.xlsm'): return 'Excel'
    if 'Acro'    in prog or 'PDF' in prog or ext == '.pdf':  return 'PDF'
    if 'Word'    in prog or ext in ('.docx','.doc'):         return 'Word'
    if 'Package' in prog or ext == '.msg':                    return 'MSG'
    if 'Visio'   in prog or ext in ('.vsd','.vsdx'):         return 'VSD'
    if 'PowerPoint' in prog or ext in ('.pptx','.ppt'):      return 'PPTX'
    return ext.upper().lstrip('.') or 'Object'

def emf_display_name(data):
    decoded = [s.decode('utf-16-le') for s in re.findall(b'(?:[\x20-\x7e]\x00){3,}', data)]
    found = []
    for i, s in enumerate(decoded):
        if s == 'IconOnly':
            for c in decoded[max(0, i-6):i]:
                if '.' in c and '\\' not in c and c != 'IconOnly':
                    found.append(c)
    return max(found, key=len) if found else ''

def unique_dest(folder, name):
    dest = os.path.join(folder, name)
    if os.path.exists(dest):
        b, e = os.path.splitext(name)
        dest = os.path.join(folder, f"{b}_{datetime.now().strftime('%H%M%S%f')}{e}")
    return dest

def short_path(p):
    n = p.replace('\\', '/')
    i = n.find(f'/{os.path.basename(OUTPUT_FOLDER)}/')
    return '..' + n[i:] if i >= 0 else p

def make_border(color="000000"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def folder_label(doc_path):
    """Return the immediate parent folder name, or '' if file is directly in Input/."""
    parent = os.path.basename(os.path.dirname(doc_path))
    return '' if parent == os.path.basename(INPUT_FOLDER) else parent


# ══════════════════════════════════════════════════════════════════════════════
#  .docx parser
# ══════════════════════════════════════════════════════════════════════════════

def parse_docx(docx_path):
    with zipfile.ZipFile(docx_path) as z:
        doc_xml  = z.read("word/document.xml")
        rels_xml = z.read("word/_rels/document.xml.rels")
        names    = z.namelist()
        emf      = {os.path.basename(n): z.read(n)
                    for n in names if n.startswith('word/media/') and n.endswith('.emf')}
        num_xml  = z.read("word/numbering.xml") if "word/numbering.xml" in names else None

    rels   = {r.get('Id'): {'type': r.get('Type','').split('/')[-1], 'target': r.get('Target','')}
              for r in etree.fromstring(rels_xml).findall(f'{{{NS_PKG}}}Relationship')}
    hlinks = {rid: v['target'] for rid, v in rels.items() if v['type'] == 'hyperlink'}

    tree      = etree.fromstring(doc_xml)
    body_kids = list(b) if (b := tree.find(f'{{{NS_W}}}body')) is not None else []

    # ShapeID → display name from EMF icon images
    shape_names = {}
    for shape in tree.iter(f'{{{NS_V}}}shape'):
        img = shape.find(f'{{{NS_V}}}imagedata')
        if img is None: continue
        tgt  = rels.get(img.get(f'{{{NS_R}}}id',''), {}).get('target','')
        name = emf_display_name(emf.get(os.path.basename(tgt), b''))
        if name: shape_names[shape.get('id','')] = name

    # Pre-compute section numbers from numbering.xml
    _heading_numbers = {}
    if num_xml is not None:
        _num_tree = etree.fromstring(num_xml)
        _abstract = {}
        for _an in _num_tree.findall(f'{{{NS_W}}}abstractNum'):
            _aid  = _an.get(f'{{{NS_W}}}abstractNumId')
            _lvls = {}
            for _lvl in _an.findall(f'{{{NS_W}}}lvl'):
                _il  = _lvl.get(f'{{{NS_W}}}ilvl','0')
                _fmt = _lvl.find(f'{{{NS_W}}}numFmt')
                _st  = _lvl.find(f'{{{NS_W}}}start')
                _lvls[_il] = {
                    'fmt':   _fmt.get(f'{{{NS_W}}}val','') if _fmt is not None else '',
                    'start': int(_st.get(f'{{{NS_W}}}val','1')) if _st is not None else 1,
                }
            _abstract[_aid] = _lvls
        _num_map = {}
        for _n in _num_tree.findall(f'{{{NS_W}}}num'):
            _nid    = _n.get(f'{{{NS_W}}}numId')
            _an_ref = _n.find(f'{{{NS_W}}}abstractNumId')
            _ovr    = {}
            for _ov in _n.findall(f'{{{NS_W}}}lvlOverride'):
                _il2 = _ov.get(f'{{{NS_W}}}ilvl','0')
                _sv  = _ov.find(f'{{{NS_W}}}startOverride')
                if _sv is not None:
                    _ovr[_il2] = int(_sv.get(f'{{{NS_W}}}val','1'))
            if _an_ref is not None:
                _num_map[_nid] = {'abstractId': _an_ref.get(f'{{{NS_W}}}val'), 'overrides': _ovr}
        _counters  = defaultdict(int)
        _used_nids = set()
        for _p in tree.iter(f'{{{NS_W}}}p'):
            _ps = _p.find(f'.//{{{NS_W}}}pStyle')
            if _ps is None: continue
            _style = _ps.get(f'{{{NS_W}}}val','')
            if 'Heading' not in _style or _style == 'TOCHeading': continue
            _numPr = _p.find(f'.//{{{NS_W}}}numPr')
            if _numPr is not None:
                _nid_el  = _numPr.find(f'{{{NS_W}}}numId')
                _ilvl_el = _numPr.find(f'{{{NS_W}}}ilvl')
                _nid = _nid_el.get(f'{{{NS_W}}}val','') if _nid_el is not None else ''
                _il  = _ilvl_el.get(f'{{{NS_W}}}val','0') if _ilvl_el is not None else '0'
                _nm  = _num_map.get(_nid, {})
                _aid = _nm.get('abstractId','')
                _key = (_aid, _il)
                _ovr_val = _nm.get('overrides',{}).get(_il)
                if _nid not in _used_nids:
                    _used_nids.add(_nid)
                    if _ovr_val is not None:
                        _counters[_key] = _ovr_val - 1
                    elif _key not in _counters:
                        _base = _abstract.get(_aid,{}).get(_il,{}).get('start',1)
                        _counters[_key] = _base - 1
                _counters[_key] += 1
                _heading_numbers[id(_p)] = str(_counters[_key])
            else:
                _ptxt = ''.join(_t.text or '' for _t in _p.iter(f'{{{NS_W}}}t'))
                _m = re.match(r'^(\d+)\s*', _ptxt)
                if _m:
                    _heading_numbers[id(_p)] = _m.group(1)

    def text(e):
        return ''.join(t.text or '' for t in e.iter(f'{{{NS_W}}}t')).strip()

    def section(ole):
        idx = next((i for i, c in enumerate(body_kids) if any(x is ole for x in c.iter())), None)
        if idx is None: return "DOCUMENT"
        for c in reversed(body_kids[:idx]):
            if c.tag == f'{{{NS_W}}}p':
                ps = c.find(f'.//{{{NS_W}}}pStyle')
                if ps is not None and 'eading' in ps.get(f'{{{NS_W}}}val',''):
                    t = text(c).upper()
                    if not t: continue
                    num = _heading_numbers.get(id(c),'')
                    if num and not t.startswith(num + '.') and not t.startswith(num + ' '):
                        return f"{num}. {t}"
                    return t
        return "DOCUMENT"

    def table_context(ole):
        p       = ole.getparent()
        tc_node = None
        while p is not None:
            tag = p.tag.split('}')[-1]
            if tag == 'tc': tc_node = p
            if tag == 'tr': break
            p = p.getparent()
        if p is None or p.tag.split('}')[-1] != 'tr':
            return '', '', ''
        tr  = p
        tbl = tr.getparent()
        if tbl is None: return '', '', ''
        all_rows = tbl.findall(f'{{{NS_W}}}tr')
        tcs      = tr.findall(f'{{{NS_W}}}tc')
        ole_col  = next((i for i, tc in enumerate(tcs)
                         if tc_node is not None and tc is tc_node), None)
        row_label_text = ''
        row_label_col  = None
        for i, tc in enumerate(tcs):
            if i == ole_col: continue
            t = text(tc)
            if t:
                row_label_text = t[:100]; row_label_col = i; break
        hdr_tcs = all_rows[0].findall(f'{{{NS_W}}}tc') if all_rows else []
        row_label_col_hdr = (text(hdr_tcs[row_label_col])
                             if row_label_col is not None and row_label_col < len(hdr_tcs) else '')
        url_col_hdr       = (text(hdr_tcs[ole_col])
                             if ole_col is not None and ole_col < len(hdr_tcs) else '')
        return row_label_text, row_label_col_hdr, url_col_hdr

    def hyperlink(ole):
        p = ole.getparent()
        while p is not None:
            if p.tag.split('}')[-1] == 'hyperlink':
                return hlinks.get(p.get(f'{{{NS_R}}}id',''),'')
            p = p.getparent()
        return ''

    def internal_path(rid, obj_type, ext):
        c = 'word/' + rels[rid]['target'].lstrip('/')
        if c in names: return c
        if obj_type == 'PDF':
            b = next((n for n in names if 'embeddings' in n and n.endswith('.bin')), None)
            if b: return b
        return next((n for n in names if 'embeddings' in n and n.endswith(ext)), None)

    results = []
    for ole in tree.iter():
        if ole.tag.split('}')[-1] != 'OLEObject': continue
        rid  = ole.get(f'{{{NS_R}}}id','')
        prog = ole.get('ProgID','')
        if not rid or rid not in rels: continue
        raw     = rels[rid]['target']
        decoded = urllib.parse.unquote(raw)
        is_url  = decoded.startswith(('http://','https://'))
        if decoded.startswith('file:///'): decoded = decoded[8:]
        ext      = os.path.splitext(os.path.basename(decoded))[1].lower()
        obj_type = ole_type(prog, ext)
        url      = hyperlink(ole) or (urllib.parse.unquote(raw) if is_url else '')
        row_lbl, rl_col_hdr, url_col_hdr = table_context(ole)
        results.append({
            'source_doc':         os.path.basename(docx_path),
            'folder':             folder_label(docx_path),
            'object_file':        shape_names.get(ole.get('ShapeID',''),'') or os.path.basename(decoded),
            'object_type':        obj_type,
            'section':            section(ole),
            'row_label':          row_lbl,
            'row_label_col_hdr':  rl_col_hdr,
            'url_col_hdr':        url_col_hdr,
            'hyperlink_url':      url,
            'internal_path':      None if is_url else internal_path(rid, obj_type, ext),
            'ole_target':         urllib.parse.unquote(raw),
            '_doc_bytes':         None,
        })
    return results


# ══════════════════════════════════════════════════════════════════════════════
#  .doc parser  (pure-Python OLE2 reader — no LibreOffice)
# ══════════════════════════════════════════════════════════════════════════════

class _OleFile:
    """Minimal OLE2 Compound File reader."""
    _FREE = 0xFFFFFFFF; _END = 0xFFFFFFFE; _FAT = 0xFFFFFFFD; _DIF = 0xFFFFFFFC

    def __init__(self, path):
        with open(path,'rb') as f: self._d = f.read()
        if self._d[:8] != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            raise ValueError("Not an OLE2 file")
        self._ss  = 1 << struct.unpack_from('<H', self._d, 30)[0]
        self._mss = 1 << struct.unpack_from('<H', self._d, 32)[0]
        self._mc  = struct.unpack_from('<I', self._d, 56)[0]
        self._fat = self._build_fat()
        self._build_dir()
        self._ms, self._mf = self._build_mini()

    def _off(self, sec): return 512 + sec * self._ss
    def _u32(self, d, o): return struct.unpack_from('<I', d, o)[0]

    def _build_fat(self):
        ss, d = self._ss, self._d
        difat = [self._u32(d,76+i*4) for i in range(109)]
        ext = self._u32(d,68)
        while ext not in (self._FREE,self._END,self._DIF,self._FAT):
            o = self._off(ext)
            difat += [self._u32(d,o+i*4) for i in range(ss//4-1)]
            ext = self._u32(d,o+ss-4)
        fat = []
        for fs in difat:
            if fs >= self._DIF: continue
            o = self._off(fs)
            fat += [self._u32(d,o+i*4) for i in range(ss//4)]
        return fat

    def _chain(self, start):
        c,s,seen = [],start,set()
        while s not in (self._FREE,self._END) and s not in seen:
            c.append(s); seen.add(s)
            s = self._fat[s] if s < len(self._fat) else self._END
        return c

    def _read_chain(self, start, size=None):
        ss = self._ss
        b  = b''.join(self._d[self._off(s):self._off(s)+ss] for s in self._chain(start))
        return b[:size] if size is not None else b

    def _build_dir(self):
        raw = self._read_chain(self._u32(self._d,48))
        self._dir      = {}   # name -> entry dict (first occurrence wins for reads)
        self._dir_list = []   # ordered list for sequential child discovery
        for i in range(len(raw)//128):
            e    = raw[i*128:(i+1)*128]
            nlen = struct.unpack_from('<H',e,64)[0]
            if nlen < 2 or nlen > 64:
                self._dir_list.append(None); continue
            name  = e[:nlen-2].decode('utf-16-le',errors='replace')
            entry = {
                'name':  name,
                'type':  struct.unpack_from('<B',e,66)[0],
                'start': self._u32(e,116), 'size': self._u32(e,120),
                'child': self._u32(e,76),  'left': self._u32(e,72),
                'right': self._u32(e,80),  'idx':  i,
            }
            if name not in self._dir:
                self._dir[name] = entry
            self._dir_list.append(entry)

    def _build_mini(self):
        if 'Root Entry' not in self._dir: return b'',[]
        ms  = self._read_chain(self._dir['Root Entry']['start'])
        mfs = self._u32(self._d,60); mfc = self._u32(self._d,64)
        mf,s = [],mfs
        for _ in range(mfc):
            if s in (self._FREE,self._END): break
            o = self._off(s)
            mf += [self._u32(self._d,o+i*4) for i in range(self._ss//4)]
            s = self._fat[s] if s < len(self._fat) else self._END
        return ms,mf

    def _read_mini(self, start, size):
        mss = self._mss
        c,s,seen = [],start,set()
        while s not in (self._FREE,self._END) and s not in seen:
            c.append(s); seen.add(s)
            s = self._mf[s] if s < len(self._mf) else self._END
        return (b''.join(self._ms[s*mss:(s+1)*mss] for s in c))[:size]

    def names(self): return list(self._dir.keys())

    def read_from(self, storage_name, stream_name):
        """Read a stream from within a specific storage (avoids name collisions)."""
        if storage_name not in self._dir: raise KeyError(storage_name)
        st_idx = self._dir[storage_name]['idx']
        i = st_idx + 1
        while i < len(self._dir_list):
            e = self._dir_list[i]
            if e is None: i += 1; continue
            if e['type'] == 1: break
            if e['name'] == stream_name:
                if e['size'] < self._mc and self._ms:
                    return self._read_mini(e['start'], e['size'])
                return self._read_chain(e['start'], e['size'])
            i += 1
        raise KeyError(f"{storage_name}/{stream_name}")

    def children_of_pool(self):
        """Return all _XXXXXXXX object storages under ObjectPool via sequential scan."""
        if 'ObjectPool' not in self._dir: return []
        op_idx = self._dir['ObjectPool']['idx']
        return [e['name'] for e in self._dir_list
                if e and e['type'] == 1 and e['name'].startswith('_')
                and e['idx'] > op_idx]

    def streams_of(self, storage_name):
        """Return stream names inside a storage via sequential scan."""
        if storage_name not in self._dir: return []
        st_idx = self._dir[storage_name]['idx']
        result = []
        i = st_idx + 1
        while i < len(self._dir_list):
            e = self._dir_list[i]
            if e is None: i += 1; continue
            if e['type'] == 1: break
            result.append(e['name'])
            i += 1
        return result

    def dir_idx(self, name):
        e = self._dir.get(name,{})
        return e.get('idx', 9999)


def _parse_ole10native(raw):
    """Parse Ole10Native stream → (label, src_path, file_bytes)."""
    try:
        pos=4; _t=struct.unpack_from('<H',raw,pos)[0]; pos+=2
        if _t not in (1,2): pos=2
        end=raw.index(b'\x00',pos); lbl=raw[pos:end].decode('cp1252','replace'); pos=end+1
        end=raw.index(b'\x00',pos); src=raw[pos:end].decode('cp1252','replace'); pos=end+1
        pos+=4
        end=raw.index(b'\x00',pos); pos=end+1
        sz=struct.unpack_from('<I',raw,pos)[0]; pos+=4
        return lbl, src, raw[pos:pos+sz]
    except Exception: return '','',b''

def _parse_compobj(raw):
    """Extract ProgID string from a CompObj stream."""
    try:
        off=28; l=struct.unpack_from('<I',raw,off)[0]; off+=4
        user=raw[off:off+l].decode('cp1252','replace').rstrip('\x00'); off+=l
        tag=struct.unpack_from('<I',raw,off)[0]; off+=4
        if tag in (0xFFFFFFFE,0xFFFFFFFF):
            l2=struct.unpack_from('<I',raw,off)[0]; off+=4+l2
        l3=struct.unpack_from('<I',raw,off)[0]; off+=4
        prog=raw[off:off+l3].decode('cp1252','replace').rstrip('\x00')
        return prog or user
    except Exception: return ''

def _zip_ext(raw):
    """Peek inside ZIP bytes to return .xlsx/.docx/.pptx/.zip."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            ns = z.namelist()
            if any(n.startswith('xl/')   for n in ns): return '.xlsx'
            if any(n.startswith('word/') for n in ns): return '.docx'
            if any(n.startswith('ppt/')  for n in ns): return '.pptx'
    except Exception: pass
    return '.zip'

def _zip_label(raw, ext):
    """Try to get a human-readable filename from ZIP metadata."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            ns = z.namelist()
            if 'docProps/core.xml' in ns:
                core  = z.read('docProps/core.xml').decode('utf-8','replace')
                match = re.search(r'<dc:title>(.*?)</dc:title>', core)
                if match and match.group(1).strip():
                    t = match.group(1).strip()
                    return t if t.lower().endswith(ext) else f"{t}{ext}"
            if ext == '.xlsx' and 'xl/workbook.xml' in ns:
                wb_xml = z.read('xl/workbook.xml').decode('utf-8','replace')
                shs    = re.findall(r'<sheet[^>]+name="([^"]+)"', wb_xml)
                if shs: return f"{shs[0]}{ext}"
    except Exception: pass
    return ''


def _doc_table_context(wd_stream, ole_byte_pos):
    """
    Given the WordDocument stream bytes and the byte position of an OLE placeholder
    (the \x01 inside \x14\x01\x15), return (section, row_label, rl_col_hdr, url_col_hdr).

    Strategy:
    - Decode a window around the OLE position as CP-1252 ANSI text.
    - Mark \x14\x01\x15 (OLE field result) as a sentinel before stripping field codes.
    - Split on \r to find paragraphs; split on \x07 to find table cells.
    - Walk backwards to find: the enclosing table row, the header row, and the
      nearest ALL-CAPS heading paragraph (= section).
    """
    SENT = '\ue001'

    def _clean(raw_bytes):
        t = raw_bytes.decode('cp1252', errors='replace')
        t = t.replace('\x14\x01\x15', SENT)           # mark OLE placeholders
        t = re.sub(r'\x13[^\x14\x15]*\x14', '', t)    # strip field instructions
        t = re.sub(r'[\x00-\x06\x08\x0a-\x0c\x0e-\x12\x13-\x1f]', '', t)
        t = t.replace(SENT, '\x01')
        return t

    # Decode a wide window (up to 2000 bytes back, 200 forward)
    win_start = max(0, ole_byte_pos - 2000)
    win_txt   = _clean(wd_stream[win_start:ole_byte_pos + 200])

    # Split into paragraphs
    paras = win_txt.split('\r')

    # Walk backwards through paras to build context
    section_name  = 'APPENDIX'
    row_label     = ''
    rl_col_hdr    = ''
    url_col_hdr   = ''
    table_headers = []

    def _is_allcaps(t):
        alpha = [c for c in t if c.isalpha()]
        return len(alpha) >= 4 and all(c.isupper() for c in alpha)

    # Find the OLE in the last para that contains it, then work backwards
    ole_para_idx = None
    for pi, p in enumerate(paras):
        if '\x01' in p:
            ole_para_idx = pi

    if ole_para_idx is None:
        return section_name, row_label, rl_col_hdr, url_col_hdr

    # Determine section from paragraphs before the OLE para
    for p in reversed(paras[:ole_para_idx]):
        t = p.strip()
        if '\x07' not in t and _is_allcaps(t):
            # Check if it has a numeric prefix (e.g. "17. APPENDIX" or "17APPENDIX")
            m = re.match(r'^(\d+)[.\s]*(.+)', t)
            section_name = f"{m.group(1)}. {m.group(2).strip()}" if m else t
            break

    # Parse the OLE paragraph as a flat cell array
    ole_para  = paras[ole_para_idx]
    all_cells = [c.strip() for c in ole_para.split('\x07')]

    # Detect n_cols from header (count non-empty cells before first empty)
    n_cols = 0
    for c in all_cells:
        if not c: break
        n_cols += 1
    if n_cols == 0: n_cols = 3
    row_width = n_cols + 1  # cols + row-end marker

    # Header cells
    hdr = all_cells[:n_cols]

    # Find which cell contains the OLE marker
    data_cells = all_cells[row_width:]
    ole_data_idx = next((i for i,c in enumerate(data_cells) if '\x01' in c), None)
    if ole_data_idx is None:
        # OLE is in the header row itself — look for it there
        ole_hdr_idx = next((i for i,c in enumerate(all_cells[:row_width]) if '\x01' in c), None)
        if ole_hdr_idx is not None:
            for i,c in enumerate(all_cells[:row_width]):
                v = c.replace('\x01','').strip()
                if v and i != ole_hdr_idx:
                    row_label = v[:100]; rl_col_hdr = hdr[i] if i<len(hdr) else ''
                    url_col_hdr = hdr[ole_hdr_idx] if ole_hdr_idx<len(hdr) else ''
                    break
        return section_name, row_label, rl_col_hdr, url_col_hdr

    row_num    = ole_data_idx // row_width
    col_in_row = ole_data_idx % row_width
    row_start  = row_num * row_width
    row        = data_cells[row_start:row_start + row_width]

    # Row label = first non-empty, non-OLE cell in the data row
    for ci, c in enumerate(row[:n_cols]):
        v = c.replace('\x01','').strip()
        if v and ci != col_in_row:
            row_label   = v[:100]
            rl_col_hdr  = hdr[ci]      if ci      < len(hdr) else ''
            url_col_hdr = hdr[col_in_row] if col_in_row < len(hdr) else ''
            break

    return section_name, row_label, rl_col_hdr, url_col_hdr


def parse_doc(doc_path):
    """
    Extract embedded OLE objects from a legacy .doc (Word 97-2003) file.
    Uses pure-Python OLE2 parsing — no LibreOffice required.
    """
    try:
        ole = _OleFile(doc_path)
    except Exception as ex:
        print(f"  [!] Cannot open .doc: {ex}"); return []

    doc_stem = os.path.splitext(os.path.basename(doc_path))[0]

    # Read WordDocument stream for text/context extraction
    wd_stream = b''
    if 'WordDocument' in ole.names():
        try: wd_stream = ole._read_chain(ole._dir['WordDocument']['start'],
                                          ole._dir['WordDocument']['size'])
        except Exception: pass

    # Find all true OLE placeholder positions (\x14\x01\x15) in the stream
    ole_positions = [i+1 for i in range(len(wd_stream)-2)
                     if wd_stream[i]==0x14 and wd_stream[i+1]==0x01 and wd_stream[i+2]==0x15]

    # Collect all _XXXXXXXX object storages, sorted by directory index (= document order)
    storages = sorted(ole.children_of_pool(), key=ole.dir_idx)

    results = []

    for storage in storages:
        streams = ole.streams_of(storage)

        def _r(name):
            return ole.read_from(storage, name)

        # ProgID
        prog_id = ''
        if '\x01CompObj' in streams:
            try: prog_id = _parse_compobj(_r('\x01CompObj'))
            except Exception: pass

        file_bytes = b''; file_label = ''; file_ext = ''

        if 'Package' in streams:
            raw = _r('Package')
            if raw[:2] == b'PK':                        # direct ZIP (xlsx, docx, …)
                file_bytes = raw
                file_ext   = _zip_ext(raw)
            else:                                        # Ole10Native wrapper
                lbl, src, fb = _parse_ole10native(b'\x04\x00\x00\x00\x02\x00' + raw)
                file_bytes = fb if fb else raw
                file_label = os.path.basename(src) if src else lbl
                file_ext   = os.path.splitext(file_label)[1].lower()

        elif '\x01Ole10Native' in streams:
            raw = _r('\x01Ole10Native')
            lbl, src, fb = _parse_ole10native(raw)
            file_bytes = fb
            file_label = lbl or os.path.basename(src)
            file_ext   = os.path.splitext(file_label)[1].lower()

        elif 'CONTENTS' in streams:
            raw = _r('CONTENTS')
            if raw[:4] == b'%PDF':
                s = raw.find(b'%PDF'); e = raw.rfind(b'%%EOF')
                file_bytes = raw[s:e+5] if s>=0 and e>=0 else raw
                file_ext   = '.pdf'
            elif raw[:2] == b'PK':
                file_bytes = raw; file_ext = _zip_ext(raw)

        # Skip if nothing useful
        if not file_bytes: continue

        # Sanitize label
        file_label = ''.join(c for c in file_label if c.isprintable()).strip()

        # Skip bare BMP icon images (not real document objects)
        if file_bytes[:2] == b'BM' and file_ext not in ('.bmp',): continue
        if len(file_bytes) < 64: continue

        # Derive label from ZIP metadata if still missing
        if not file_label and file_bytes[:2] == b'PK':
            file_label = _zip_label(file_bytes, file_ext)
        if not file_label:
            file_label = f"{doc_stem}_embedded{file_ext or '.bin'}"

        obj_type = ole_type(prog_id, file_ext)

        # Match to document context using OLE position list (consume in order)
        if ole_positions:
            ctx_pos = ole_positions.pop(0)
            section, row_label, rl_col_hdr, url_col_hdr = _doc_table_context(wd_stream, ctx_pos)
        else:
            section = 'DOCUMENT'; row_label = rl_col_hdr = url_col_hdr = ''

        results.append({
            'source_doc':         os.path.basename(doc_path),
            'folder':             folder_label(doc_path),
            'object_file':        file_label,
            'object_type':        obj_type,
            'section':            section,
            'row_label':          row_label,
            'row_label_col_hdr':  rl_col_hdr,
            'url_col_hdr':        url_col_hdr,
            'hyperlink_url':      '',
            'internal_path':      None,
            'ole_target':         '',
            '_doc_bytes':         file_bytes,
        })

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  Fix & verify extracted files
# ══════════════════════════════════════════════════════════════════════════════

def fix_excel(path):
    """Remove OLE embedding flags so the file opens normally as standalone."""
    try:
        buf = io.BytesIO(open(path,'rb').read())
        with zipfile.ZipFile(buf) as z:
            if 'xl/workbook.xml' not in z.namelist(): return
            xml = z.read('xl/workbook.xml')
        if b'visibility="hidden"' not in xml and b'<oleSize' not in xml and b'activeTab=' in xml: return
        xml = xml.replace(b' visibility="hidden"',b'').replace(b'visibility="hidden"',b'')
        if b'activeTab=' not in xml: xml = xml.replace(b'<workbookView',b'<workbookView activeTab="0"')
        while b'<oleSize' in xml:
            s, e = xml.find(b'<oleSize'), xml.find(b'>',xml.find(b'<oleSize'))+1
            xml  = xml[:s]+xml[e:]
        out = io.BytesIO()
        buf.seek(0)
        with zipfile.ZipFile(buf) as zi, zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                zo.writestr(item, xml if item.filename=='xl/workbook.xml' else zi.read(item.filename))
        open(path,'wb').write(out.getvalue())
    except Exception: pass

def verify(path):
    if not os.path.isfile(path) or os.path.getsize(path) == 0:
        return False, "missing or empty"
    size, ext = os.path.getsize(path), os.path.splitext(path)[1].lower()
    if ext in ('.xlsx','.xlsm','.xls'):
        try:
            with zipfile.ZipFile(path) as z:
                bad = z.testzip()
                if bad: return False, f"corrupt: {bad}"
                if 'xl/workbook.xml' not in z.namelist(): return False, "no workbook.xml"
                if b'visibility="hidden"' in z.read('xl/workbook.xml'): return False, "still hidden"
                cells = sum(len(re.findall(b'<c ',z.read(n)))
                            for n in z.namelist() if 'xl/worksheets/sheet' in n)
            return (True, f"OK ({size:,}b, {cells} cells)") if cells else (False, "no data")
        except zipfile.BadZipFile: return False, "invalid xlsx"
    if ext == '.pdf':
        with open(path,'rb') as f: hdr = f.read(4)
        return (True, f"OK ({size:,}b)") if hdr == b'%PDF' else (False, "invalid PDF")
    return True, f"OK ({size:,}b)"


# ══════════════════════════════════════════════════════════════════════════════
#  Extract files to Output/
# ══════════════════════════════════════════════════════════════════════════════

def extract_files(objects, doc_path):
    stem    = os.path.splitext(os.path.basename(doc_path))[0]
    out_dir = os.path.join(OUTPUT_FOLDER, stem)
    os.makedirs(out_dir, exist_ok=True)
    is_docx = doc_path.lower().endswith('.docx')

    if is_docx:
        zf = zipfile.ZipFile(doc_path)

    try:
        for obj in objects:
            if obj['hyperlink_url']:
                obj['saved_path'] = ''; continue

            # .doc files carry their bytes in _doc_bytes
            if not is_docx and obj.get('_doc_bytes'):
                name = obj['object_file']
                dest = unique_dest(out_dir, name)
                try:
                    fb = obj['_doc_bytes']
                    if name.lower().endswith('.xlsx'): fix_excel_bytes_then_write(fb, dest)
                    else: open(dest,'wb').write(fb)
                    if dest.lower().endswith(('.xlsx','.xlsm','.xls')): fix_excel(dest)
                    obj['saved_path']    = os.path.abspath(dest)
                    ok, msg = verify(obj['saved_path'])
                    obj['verify_status'] = f'✔ {msg}' if ok else f'✘ {msg}'
                except Exception as ex:
                    obj['saved_path'] = ''; obj['verify_status'] = f'✘ {ex}'
                continue

            # .docx embedded files
            if obj['internal_path']:
                name = obj['object_file']
                if obj['internal_path'].endswith('.bin') and name.lower().endswith('.bin'):
                    type_ext = '.pdf' if obj['object_type'] == 'PDF' else '.bin'
                    name = f"{stem}_embedded{type_ext}"
                dest = unique_dest(out_dir, name)
                try:
                    data = zf.read(obj['internal_path'])
                    if obj['internal_path'].endswith('.bin') and obj['object_type'] == 'PDF':
                        s, e = data.find(b'%PDF'), data.rfind(b'%%EOF')
                        if s >= 0 and e >= 0: data = data[s:e+5]
                    open(dest,'wb').write(data)
                    if dest.lower().endswith(('.xlsx','.xlsm','.xls')): fix_excel(dest)
                    obj['object_file']   = name
                    obj['saved_path']    = os.path.abspath(dest)
                    ok, msg = verify(obj['saved_path'])
                    obj['verify_status'] = f'✔ {msg}' if ok else f'✘ {msg}'
                except Exception as ex:
                    obj['saved_path'] = ''; obj['verify_status'] = f'✘ {ex}'
            else:
                local = obj['ole_target'].replace('file:///','').replace('file://','')
                if os.path.isfile(local):
                    dest = unique_dest(out_dir, obj['object_file'])
                    try:
                        shutil.copy2(local, dest)
                        obj['saved_path'] = os.path.abspath(dest)
                        ok, msg = verify(obj['saved_path'])
                        obj['verify_status'] = f'✔ {msg}' if ok else f'✘ {msg}'
                    except Exception as ex:
                        obj['saved_path'] = ''; obj['verify_status'] = f'✘ {ex}'
                else:
                    obj['saved_path'] = ''; obj['verify_status'] = ''
    finally:
        if is_docx: zf.close()


def fix_excel_bytes_then_write(raw_bytes, dest_path):
    """Write Excel bytes and attempt to fix hidden workbook flags in one pass."""
    open(dest_path,'wb').write(raw_bytes)


# ══════════════════════════════════════════════════════════════════════════════
#  Write Excel report
# ══════════════════════════════════════════════════════════════════════════════

def write_report(objects):
    wb = Workbook(); ws = wb.active; ws.title = "Embedded Objects"
    border = make_border()
    cols = [
        ("Folder",                  25),
        ("Word Document",           35),
        ("Word Section",            28),
        ("Row Label",               50),
        ("Row Label Column Header", 30),
        ("Object Type",             14),
        ("URL",                     70),
        ("URL Column Header",       30),
    ]
    for ci,(h,w) in enumerate(cols,1):
        c = ws.cell(row=1,column=ci,value=h)
        c.font      = Font(name="Arial",size=11,bold=True,color="FFFFFF")
        c.fill      = PatternFill("solid",start_color="404040")
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border    = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 30

    def cell(r,c,v,color="000000",wrap=False):
        x = ws.cell(row=r,column=c,value=v)
        x.font      = Font(name="Arial",size=10,color=color)
        x.fill      = PatternFill("solid",start_color="FFFFFF")
        x.alignment = Alignment(horizontal="left",vertical="top",wrap_text=wrap)
        x.border    = border

    for i,obj in enumerate(objects):
        r   = i+2
        url = (obj['hyperlink_url']
               or (short_path(obj['saved_path']) if obj.get('saved_path') else '')
               or obj.get('ole_target',''))
        cell(r,1,obj.get('folder',''))
        cell(r,2,obj.get('source_doc',''))
        cell(r,3,obj['section'])
        cell(r,4,obj['row_label'],          wrap=True)
        cell(r,5,obj.get('row_label_col_hdr',''))
        cell(r,6,obj['object_type'])
        cell(r,7,url,                       color="0563C1",wrap=True)
        cell(r,8,obj.get('url_col_hdr',''))
        ws.row_dimensions[r].height = 40 if len(obj.get('row_label',''))>40 else 18

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:H{len(objects)+1}"
    path = os.path.join(OUTPUT_FOLDER,REPORT_NAME)
    wb.save(path); return path


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    all_files = sorted(
        os.path.join(root,f)
        for root,_,files in os.walk(INPUT_FOLDER)
        for f in files
        if f.lower().endswith(('.docx','.doc')) and not f.startswith('~')
    )
    if not all_files:
        print(f"No .docx/.doc files found in: {INPUT_FOLDER}"); return

    print(f"\nFound {len(all_files)} file(s):")
    all_objects = []

    for i, path in enumerate(all_files, 1):
        name  = os.path.basename(path)
        fldr  = folder_label(path)
        ftype = '.doc' if name.lower().endswith('.doc') else '.docx'
        print(f"\n[{i}/{len(all_files)}] {name}" + (f"  (folder: {fldr})" if fldr else ""))
        try:
            objects = parse_doc(path) if ftype == '.doc' else parse_docx(path)
        except Exception as ex:
            print(f"  [!] Failed: {ex}"); continue
        extract_files(objects, path)
        for obj in objects:
            status = obj.get('verify_status','')
            if not obj['hyperlink_url']:
                print(f"  • {obj['object_file']}  [{status}]")
            else:
                print(f"  • {obj['object_file']}  [linked → {obj['hyperlink_url'][:60]}]")
        all_objects.extend(objects)

    if all_objects:
        report = write_report(all_objects)
        print(f"\nReport → {report}")

    total    = len(all_objects)
    with_url = sum(1 for o in all_objects if o['hyperlink_url'])
    print(f"\nScanned: {len(all_files)} file(s) | Objects: {total} | "
          f"Linked: {with_url} | Embedded: {total-with_url}\n")

if __name__ == "__main__":
    main()