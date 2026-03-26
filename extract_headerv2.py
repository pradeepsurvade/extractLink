import os, sys, zipfile, re, struct
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

W_NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

# ─── Placeholder detection ────────────────────────────────────────────────────

def is_placeholder(value):
    """
    Return True if a header value is an unfilled template placeholder rather
    than real content. Examples:
      '(This should come from the process map)'  → True
      'SOP REFERENCE NO. AND SOP NAME'           → True
      'AMEA_AER_AUS'                             → False
      'GLOBAL'                                   → False
    """
    v = value.strip()
    if not v:
        return True
    # Parenthetical instruction e.g. "(This should come from the process map)"
    if (v.startswith('(') and v.endswith(')') and
            re.search(r'should|come from|process map|refer|insert|enter|'
                      r'type|tbd|n/a|placeholder|from the', v, re.IGNORECASE)):
        return True
    # Explicit all-caps template marker phrases
    if re.search(r'SOP\s+REFERENCE\s+NO|REFERENCE\s+NO\.?\s+AND\s+(?:SOP\s+)?NAME|'
                 r'(?:INSERT|ENTER|TYPE|ADD)\s+\w+\s+HERE', v, re.IGNORECASE):
        return True
    # Standalone N/A or TBD
    if re.match(r'^(N/A|TBD|NA|TO\s+BE\s+DETERMINED)$', v, re.IGNORECASE):
        return True
    return False


def clean_pairs(pairs):
    """
    If the majority of non-empty values are placeholders, the document header
    is an unfilled template — return [] so the file appears with no header data.
    Otherwise replace individual placeholder values with empty string.
    """
    if not pairs:
        return pairs

    non_empty_values = [(label, value) for label, value in pairs if value.strip()]
    if non_empty_values:
        placeholder_count = sum(1 for _, v in non_empty_values if is_placeholder(v))
        # If more than half the filled-in values are placeholders → whole record is template
        if placeholder_count >= len(non_empty_values) / 2:
            return []

    return [(label, '' if is_placeholder(value) else value)
            for label, value in pairs]


# ─── Pure-Python .doc OLE stream reader ──────────────────────────────────────

def read_ole_stream(data, stream_name):
    sector_size      = 1 << struct.unpack_from('<H', data, 30)[0]
    mini_sector_size = 1 << struct.unpack_from('<H', data, 32)[0]
    mini_cutoff      = struct.unpack_from('<I', data, 56)[0]

    fat = []
    for i in range(109):
        sec = struct.unpack_from('<I', data, 76 + i*4)[0]
        if sec >= 0xFFFFFFFE:
            break
        offset = (sec + 1) * sector_size
        for j in range(sector_size // 4):
            fat.append(struct.unpack_from('<I', data, offset + j*4)[0])

    def get_chain(start):
        chain, sec = [], start
        while sec < 0xFFFFFFFE:
            chain.append(sec)
            sec = fat[sec] if sec < len(fat) else 0xFFFFFFFE
        return chain

    def read_sectors(start, size=-1):
        raw = b''.join(data[(s+1)*sector_size:(s+1)*sector_size+sector_size]
                       for s in get_chain(start))
        return raw if size < 0 else raw[:size]

    dir_data = read_sectors(struct.unpack_from('<I', data, 48)[0])
    streams, root_start, root_size = {}, None, None
    for i in range(len(dir_data) // 128):
        e     = dir_data[i*128:(i+1)*128]
        nl    = struct.unpack_from('<H', e, 64)[0]
        if nl == 0:
            continue
        name  = e[:nl-2].decode('utf-16-le', errors='ignore')
        etype = e[66]
        start = struct.unpack_from('<I', e, 116)[0]
        size  = struct.unpack_from('<I', e, 120)[0]
        if i == 0:
            root_start, root_size = start, size
        if etype in (1, 2):
            streams[name] = (start, size)

    if stream_name not in streams:
        return None
    start, size = streams[stream_name]

    if size < mini_cutoff and root_start is not None:
        mini_fat_start = struct.unpack_from('<I', data, 60)[0]
        mini_fat = []
        for s in get_chain(mini_fat_start):
            offset = (s + 1) * sector_size
            for j in range(sector_size // 4):
                mini_fat.append(struct.unpack_from('<I', data, offset + j*4)[0])
        root_data = read_sectors(root_start, root_size)
        result, sec = b'', start
        while sec < 0xFFFFFFFE:
            offset = sec * mini_sector_size
            result += root_data[offset:offset + mini_sector_size]
            sec = mini_fat[sec] if sec < len(mini_fat) else 0xFFFFFFFE
        return result[:size]

    return read_sectors(start, size)


# ─── .doc header extraction ───────────────────────────────────────────────────

# Noise patterns to reject labels/values that are binary garbage or body content
_BAD_LABEL = re.compile(r'[\\@\x80-\xff]|HYPERLINK|MERGEFORMAT|REF |gd[A-Z]|Ifa|OJQJ')
_BAD_VALUE = re.compile(r'\\\\|OJQJ|Ifa|gdÉ|\\\\^J')

def _parse_cell(seg):
    """
    Try to parse a \x07-delimited cell into a (label, value) pair.
    Returns None if the cell doesn't look like a clean header field.
    """
    clean = re.sub(r'[\x00-\x06\x08-\x0c\x0e-\x1f\x7f-\x9f]', '', seg).strip()
    if not clean or ':' not in clean or len(clean) > 200:
        return None
    paras = [p.strip() for p in re.split(r'[\r\n]+', clean) if p.strip()]
    if not paras:
        return None
    colon_paras = [p for p in paras if ':' in p]
    if not colon_paras:
        return None

    lp    = colon_paras[0]
    idx   = lp.index(':')
    label = lp[:idx].strip()
    if len(paras) == 1:
        value = lp[idx+1:].strip()
    else:
        value = ' '.join(p for p in paras if p != lp).strip()

    # Reject labels that are too long or contain noise patterns
    if not label or len(label) > 60 or _BAD_LABEL.search(label):
        return None
    # Reject values that are clearly binary/markup noise
    if _BAD_VALUE.search(value[:50]):
        return None

    return (label, value)


def _score_group(group):
    """
    Score a group of consecutive header cells.
    Higher score = more likely to be the SOP page header.
    Rewards labels that look like 'Organization', 'Process Map', 'SOP No'.
    """
    score = 0
    keywords = ['organ', 'process map', 'sop no', 'sop number', 'sla', 'title']
    for _, (label, _) in group:
        ll = label.lower()
        if any(k in ll for k in keywords):
            score += 2
        if len(label) < 40:
            score += 1
    return score


def extract_doc_header_fields(doc_path):
    """
    Extract page-header label/value pairs from a legacy .doc file.
    Uses pure Python OLE parsing — no LibreOffice or admin rights required.

    Approach:
      1. Read the WordDocument OLE stream and decode as cp1252.
      2. Split on \\x07 (Word table cell separator).
      3. Parse each cell into (label, value) — same logic as the .docx extractor.
      4. Filter out binary noise and body-table content using heuristics.
      5. Group consecutive valid cells; pick the group that looks most like
         a SOP page header (contains Organisation/Process Map/SOP No. labels).
      6. Deduplicate repeated header sections (odd/even/first-page headers).
    """
    try:
        with open(doc_path, 'rb') as f:
            data = f.read()
        if data[:8] != bytes.fromhex('D0CF11E0A1B11AE1'):
            return []
        wd = read_ole_stream(data, 'WordDocument')
        if wd is None:
            return []
        text = wd.decode('cp1252', errors='replace')
    except Exception:
        return []

    # Parse all cells
    segments = re.split(r'\x07', text)
    parsed   = [(i, _parse_cell(s)) for i, s in enumerate(segments)]
    parsed   = [(i, p) for i, p in parsed if p is not None]

    if not parsed:
        return []

    # Group consecutive valid cells (within 3 segment indices of each other)
    groups  = []
    current = [parsed[0]]
    for entry in parsed[1:]:
        if entry[0] - current[-1][0] <= 3:
            current.append(entry)
        else:
            if len(current) >= 2:
                groups.append(current)
            current = [entry]
    if len(current) >= 2:
        groups.append(current)

    if not groups:
        # Fallback: return all parsed pairs if no groups found
        seen, pairs = set(), []
        for _, (label, value) in parsed:
            if label not in seen:
                seen.add(label)
                pairs.append((label, value))
        return pairs

    # Pick the group that best matches a SOP page header
    best_group = max(groups, key=_score_group)

    # Deduplicate labels within the best group (same header may repeat)
    seen, pairs = set(), []
    for _, (label, value) in best_group:
        if label not in seen:
            seen.add(label)
            pairs.append((label, value))

    return pairs


# ─── .docx header extraction ─────────────────────────────────────────────────

def para_text(p):
    return re.sub(r'  +', ' ', ''.join(
        t.text for t in p.iter(f'{{{W_NS}}}t') if t.text)).strip()

def extract_docx_header_fields(docx_path):
    """Extract header fields from a .docx (ZIP/XML) file."""
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            header_files = sorted(
                n for n in z.namelist()
                if re.match(r'word/header\d+\.xml$', n)
            )
            header_xml = None
            for hf in header_files:
                content = z.read(hf)
                if b'<w:tbl' not in content:
                    continue
                root  = ET.fromstring(content)
                texts = [t.text for t in root.iter(f'{{{W_NS}}}t')
                         if t.text and t.text.strip()]
                if texts:
                    header_xml = content
                    break
            if header_xml is None:
                return []
        root = ET.fromstring(header_xml)
    except Exception as e:
        print(f'  [warn] Could not read {docx_path}: {e}')
        return []

    pairs = []
    for cell in root.iter(f'{{{W_NS}}}tc'):
        non_empty = [para_text(p) for p in cell.iter(f'{{{W_NS}}}p')]
        non_empty = [p for p in non_empty if p]
        if not non_empty:
            continue
        colon_paras = [p for p in non_empty if ':' in p]
        if not colon_paras:
            pairs.append(('', ' '.join(non_empty).strip()))
            continue
        lp = colon_paras[0]
        if len(non_empty) == 1:
            idx   = lp.index(':')
            label = lp[:idx].strip()
            value = lp[idx+1:].strip()
        else:
            label     = lp
            remaining = [p for p in non_empty if p != lp]
            value     = ' '.join(remaining).strip()
        pairs.append((label, value))
    return pairs


def extract_header_fields(file_path):
    """Dispatch to the correct extractor based on file extension."""
    if str(file_path).lower().endswith('.docx'):
        pairs = extract_docx_header_fields(file_path)
    else:
        pairs = extract_doc_header_fields(file_path)
    return clean_pairs(pairs)


# ─── Folder scan ─────────────────────────────────────────────────────────────

def scan_word_folder(folder):
    """Scan folder/subfolders for .docx and .doc files; return records and max_fields."""
    records    = []
    max_fields = 0
    folder     = Path(folder)

    all_files = [
        f for f in sorted(folder.rglob('*'))
        if f.suffix.lower() in ('.doc', '.docx') and f.is_file()
    ]

    # Prefer .docx over .doc when both exist with the same stem in the same folder
    seen = {}
    for f in all_files:
        key = (f.parent, f.stem.lower())
        if key not in seen or f.suffix.lower() == '.docx':
            seen[key] = f

    for file_path in sorted(seen.values(), key=lambda p: p.name):
        print(f'  Processing: {file_path.name}')
        pairs      = extract_header_fields(file_path)
        max_fields = max(max_fields, len(pairs))
        records.append({
            '_folder': file_path.parent.name,
            '_file':   file_path.name,
            '_pairs':  pairs,
        })

    return records, max_fields


# ─── Excel formatting helpers ─────────────────────────────────────────────────

DARK_BLUE   = 'FF1F4E79'
MED_BLUE    = 'FF2E75B6'
LIGHT_GREY  = 'FFF2F2F2'
WHITE_FONT  = 'FFFFFFFF'
DARK_FONT   = 'FF1F4E79'
BORDER_DARK = 'FF4472C4'
BORDER_GREY = 'FFBFBFBF'

def thin_border(color):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border(color):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border_no_top(color):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, bottom=s)

def apply_main_header(cell, value):
    cell.value     = value
    cell.font      = Font(bold=True, color=WHITE_FONT, name='Calibri', size=11)
    cell.fill      = PatternFill('solid', fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border(BORDER_GREY)

def apply_sub_header(cell, value):
    cell.value     = value
    cell.font      = Font(bold=True, color=WHITE_FONT, name='Calibri', size=11)
    cell.fill      = PatternFill('solid', fgColor=MED_BLUE)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border    = thin_border(BORDER_GREY)

def apply_folder_file(cell, value, is_top_row):
    cell.value     = value
    cell.font      = Font(bold=True, color=DARK_FONT, name='Calibri', size=11)
    cell.fill      = PatternFill('solid', fgColor=LIGHT_GREY)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border    = (medium_border(BORDER_DARK) if is_top_row
                      else medium_border_no_top(BORDER_DARK))

def apply_data(cell, value):
    cell.value     = value
    cell.font      = Font(name='Calibri', size=11)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border    = thin_border(BORDER_GREY)


# ─── Column label configuration (edit here to rename output columns) ─────────
# These are the only three fixed column names in the output.
# Col 1: the folder containing each Word file
# Col 2: the Word filename
# Col 3: merged heading that spans all dynamic header-field columns
COL_FOLDER  = 'Folder Name'
COL_FILE    = 'File Name'
COL_HEADER  = 'Header Content'

# ─── Build Excel ──────────────────────────────────────────────────────────────

def build_excel(records, max_fields, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Tables'

    apply_main_header(ws.cell(1, 1), COL_FOLDER)
    apply_main_header(ws.cell(1, 2), COL_FILE)
    apply_main_header(ws.cell(1, 3), COL_HEADER)
    if max_fields > 1:
        ws.merge_cells(start_row=1, start_column=3,
                       end_row=1,   end_column=2 + max_fields)
    ws.row_dimensions[1].height = 28

    excel_row = 2
    for rec in records:
        pairs = rec['_pairs']

        for ci, (label, _) in enumerate(pairs, start=3):
            apply_sub_header(ws.cell(excel_row, ci), label)
        for ci in range(3 + len(pairs), 3 + max_fields):
            apply_sub_header(ws.cell(excel_row, ci), '')
        apply_folder_file(ws.cell(excel_row, 1), rec['_folder'], is_top_row=True)
        apply_folder_file(ws.cell(excel_row, 2), rec['_file'],   is_top_row=True)
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        for ci, (_, value) in enumerate(pairs, start=3):
            apply_data(ws.cell(excel_row, ci), value)
        for ci in range(3 + len(pairs), 3 + max_fields):
            apply_data(ws.cell(excel_row, ci), '')
        apply_folder_file(ws.cell(excel_row, 1), None, is_top_row=False)
        apply_folder_file(ws.cell(excel_row, 2), None, is_top_row=False)
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        ws.merge_cells(start_row=excel_row-2, start_column=1,
                       end_row=excel_row-1,   end_column=1)
        ws.merge_cells(start_row=excel_row-2, start_column=2,
                       end_row=excel_row-1,   end_column=2)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 42
    for i in range(max_fields):
        ws.column_dimensions[get_column_letter(3 + i)].width = 32

    wb.save(output_path)
    print(f'\nSaved: {output_path}')
    print(f'  Records   : {len(records)}')
    print(f'  Max fields: {max_fields}')
    for r in records:
        print(f'  {r["_file"]}:')
        for label, value in r['_pairs']:
            print(f'    {label!r:45s} → {value!r}')
        if not r['_pairs']:
            print('    (no header fields extracted)')


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    INPUT_FOLDER = Path(__file__).parent / "input"
    OUTPUT_PATH  = Path(__file__).parent / "output/header_extract_formatted.xlsx"

    print(f'Scanning: {INPUT_FOLDER}')
    if not INPUT_FOLDER.exists():
        print(f'[ERROR] Input folder not found: {INPUT_FOLDER}')
        sys.exit(1)

    records, max_fields = scan_word_folder(INPUT_FOLDER)
    if not records:
        print('No .docx or .doc files found.')
    else:
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        build_excel(records, max_fields, OUTPUT_PATH)