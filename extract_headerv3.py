import sys, zipfile, re, struct
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

W_NS       = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
COL_FOLDER = 'Folder Name'
COL_FILE   = 'File Name'
COL_HEADER = 'Header Content'

# ─── Compiled patterns ────────────────────────────────────────────────────────

_RE_CTRL        = re.compile(r'[\x00-\x06\x08-\x0c\x0e-\x1f\x7f-\x9f]')
_RE_SPACES      = re.compile(r'  +')
_RE_PLACEHOLDER = re.compile(
    r'SOP\s+REFERENCE\s+NO|REFERENCE\s+NO\.?\s+AND\s+(?:SOP\s+)?NAME|'
    r'(?:INSERT|ENTER|TYPE|ADD)\s+\w+\s+HERE|^(N/A|TBD|NA|TO\s+BE\s+DETERMINED)$',
    re.IGNORECASE
)
_RE_PAREN_HINT  = re.compile(
    r'should|come from|process map|refer|insert|enter|type|tbd|n/a|placeholder|from the',
    re.IGNORECASE
)
_RE_BAD_LABEL   = re.compile(r'[\\@\x80-\xff]|HYPERLINK|MERGEFORMAT|REF |gd[A-Z]|Ifa|OJQJ')
_RE_BAD_VALUE   = re.compile(r'\\\\|OJQJ|Ifa|gdE|\\\\^J')
_HEADER_FILE_RE = re.compile(r'word/header\d+\.xml$')
_PARA_SEP       = re.compile(r'[\r\n]+')
_SCORE_KW       = re.compile(r'organ|process map|sop no|sop number|sla|title', re.IGNORECASE)

# ─── Shared cell parser ───────────────────────────────────────────────────────

def _parse_paras(paras):
    """Convert paragraph list from one table cell into (label, value) or None."""
    colon = [p for p in paras if ':' in p]
    if not colon:
        return ('', ' '.join(paras)) if paras else None
    lp  = colon[0]
    idx = lp.index(':')
    if len(paras) == 1:
        return lp[:idx].strip(), lp[idx+1:].strip()
    return lp, ' '.join(p for p in paras if p != lp).strip()

# ─── Placeholder detection ────────────────────────────────────────────────────

def _is_placeholder(v):
    v = v.strip()
    if not v:
        return True
    if v.startswith('(') and v.endswith(')') and _RE_PAREN_HINT.search(v):
        return True
    return bool(_RE_PLACEHOLDER.search(v))


def _clean_pairs(pairs):
    """Return [] if majority of values are placeholders; else blank placeholders."""
    filled = [(l, v) for l, v in pairs if v.strip()]
    if filled and sum(_is_placeholder(v) for _, v in filled) >= len(filled) / 2:
        return []
    return [(l, '' if _is_placeholder(v) else v) for l, v in pairs]

# ─── OLE stream reader (.doc) ─────────────────────────────────────────────────

def _read_ole_stream(data, stream_name):
    ss   = 1 << struct.unpack_from('<H', data, 30)[0]
    mss  = 1 << struct.unpack_from('<H', data, 32)[0]
    mcut = struct.unpack_from('<I', data, 56)[0]

    fat = []
    for i in range(109):
        sec = struct.unpack_from('<I', data, 76 + i*4)[0]
        if sec >= 0xFFFFFFFE:
            break
        fat.extend(struct.unpack_from(f'<{ss//4}I', data, (sec+1)*ss))

    def _chain(start):
        sec = start
        while sec < 0xFFFFFFFE:
            yield sec
            sec = fat[sec] if sec < len(fat) else 0xFFFFFFFE

    def _read(start, size=-1):
        raw = b''.join(data[(s+1)*ss:(s+1)*ss+ss] for s in _chain(start))
        return raw if size < 0 else raw[:size]

    dir_data = _read(struct.unpack_from('<I', data, 48)[0])
    streams, root_start, root_size = {}, None, None
    for i in range(len(dir_data) // 128):
        e  = dir_data[i*128:(i+1)*128]
        nl = struct.unpack_from('<H', e, 64)[0]
        if not nl:
            continue
        name  = e[:nl-2].decode('utf-16-le', errors='ignore')
        start = struct.unpack_from('<I', e, 116)[0]
        size  = struct.unpack_from('<I', e, 120)[0]
        if i == 0:
            root_start, root_size = start, size
        if e[66] in (1, 2):
            streams[name] = (start, size)

    if stream_name not in streams:
        return None
    start, size = streams[stream_name]

    if size < mcut and root_start is not None:
        mini_fat = []
        for s in _chain(struct.unpack_from('<I', data, 60)[0]):
            mini_fat.extend(struct.unpack_from(f'<{ss//4}I', data, (s+1)*ss))
        root   = _read(root_start, root_size)
        result, sec = b'', start
        while sec < 0xFFFFFFFE:
            result += root[sec*mss:(sec+1)*mss]
            sec = mini_fat[sec] if sec < len(mini_fat) else 0xFFFFFFFE
        return result[:size]

    return _read(start, size)

# ─── .doc header extraction ───────────────────────────────────────────────────

def _extract_doc(doc_path):
    try:
        data = Path(doc_path).read_bytes()
        if data[:8] != bytes.fromhex('D0CF11E0A1B11AE1'):
            return []
        wd = _read_ole_stream(data, 'WordDocument')
        text = wd.decode('cp1252', errors='replace') if wd else None
        if not text:
            return []
    except Exception:
        return []

    def _parse_seg(seg):
        clean = _RE_CTRL.sub('', seg).strip()
        if not clean or ':' not in clean or len(clean) > 200:
            return None
        paras  = [p.strip() for p in _PARA_SEP.split(clean) if p.strip()]
        result = _parse_paras(paras)
        if not result:
            return None
        label, value = result
        if not label or len(label) > 60 or _RE_BAD_LABEL.search(label):
            return None
        if _RE_BAD_VALUE.search(value[:50]):
            return None
        return label, value

    parsed = [(i, p) for i, s in enumerate(re.split(r'\x07', text))
              if (p := _parse_seg(s)) is not None]
    if not parsed:
        return []

    # Group consecutive valid cells; pick the group most like a SOP page header
    groups, current = [], [parsed[0]]
    for entry in parsed[1:]:
        if entry[0] - current[-1][0] <= 3:
            current.append(entry)
        else:
            if len(current) >= 2:
                groups.append(current)
            current = [entry]
    if len(current) >= 2:
        groups.append(current)

    source = (max(groups, key=lambda g: sum(
        2 * bool(_SCORE_KW.search(lbl)) + (len(lbl) < 40)
        for _, (lbl, _) in g
    )) if groups else parsed)

    seen, pairs = set(), []
    for _, (label, value) in source:
        key = label.rstrip(':').strip().lower()
        if key not in seen:
            seen.add(key)
            pairs.append((label, value))
    return pairs

# ─── .docx header extraction ─────────────────────────────────────────────────

def _extract_docx(docx_path):
    try:
        with zipfile.ZipFile(docx_path) as z:
            header_xml = None
            for hf in sorted(z.namelist()):
                if not _HEADER_FILE_RE.match(hf):
                    continue
                content = z.read(hf)
                if b'<w:tbl' not in content:
                    continue
                root = ET.fromstring(content)
                if any(t.text and t.text.strip() for t in root.iter(f'{{{W_NS}}}t')):
                    header_xml = content
                    break
        if header_xml is None:
            return []
    except Exception:
        return []

    pairs = []
    for cell in ET.fromstring(header_xml).iter(f'{{{W_NS}}}tc'):
        paras = [_RE_SPACES.sub(' ', ''.join(
                     t.text for t in p.iter(f'{{{W_NS}}}t') if t.text)).strip()
                 for p in cell.iter(f'{{{W_NS}}}p')]
        result = _parse_paras([p for p in paras if p])
        if result:
            pairs.append(result)
    return pairs

# ─── Unified extractor ────────────────────────────────────────────────────────

def extract_header_fields(file_path):
    fn = _extract_docx if str(file_path).lower().endswith('.docx') else _extract_doc
    return _clean_pairs(fn(file_path))

# ─── Folder scan ─────────────────────────────────────────────────────────────

def scan_word_folder(folder):
    all_files = [f for f in sorted(Path(folder).rglob('*'))
                 if f.suffix.lower() in ('.doc', '.docx') and f.is_file()]

    seen = {}
    for f in all_files:
        key = (f.parent, f.stem.lower())
        if key not in seen or f.suffix.lower() == '.docx':
            seen[key] = f

    records, max_fields = [], 0
    for fp in sorted(seen.values(), key=lambda p: p.name):
        print(f'  Processing: {fp.name}')
        pairs      = extract_header_fields(fp)
        max_fields = max(max_fields, len(pairs))
        records.append({'_folder': fp.parent.name, '_file': fp.name, '_pairs': pairs})

    return records, max_fields

# ─── Excel styling ────────────────────────────────────────────────────────────

_DARK_BLUE  = 'FF1F4E79'
_MED_BLUE   = 'FF2E75B6'
_LIGHT_GREY = 'FFF2F2F2'
_WHITE      = 'FFFFFFFF'
_DARK_FONT  = 'FF1F4E79'
_GREY_SIDE  = 'FFBFBFBF'
_BLUE_SIDE  = 'FF4472C4'

def _border(color, top=True):
    thin   = color == _GREY_SIDE
    s      = Side(style='thin' if thin else 'medium', color=color)
    top_s  = s if top else Side()
    return Border(left=s, right=s, top=top_s, bottom=s)

_B_THIN   = _border(_GREY_SIDE)
_B_MED    = _border(_BLUE_SIDE)
_B_MED_NT = _border(_BLUE_SIDE, top=False)

def _style(cell, value, bold, font_color, fill, h, v, border):
    cell.value     = value
    cell.font      = Font(bold=bold, color=font_color, name='Calibri', size=11)
    cell.fill      = PatternFill('solid', fgColor=fill)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=True)
    cell.border    = border

_BLACK     = 'FF000000'

def _main_hdr(c, v):  _style(c, v, True,  _WHITE,     _DARK_BLUE,  'center', 'center', _B_THIN)
def _sub_hdr(c, v):   _style(c, v, True,  _WHITE,     _MED_BLUE,   'left',   'center', _B_THIN)
def _ff(c, v, top):   _style(c, v, True,  _DARK_FONT, _LIGHT_GREY, 'left',   'top',    _B_MED if top else _B_MED_NT)
def _data(c, v):      _style(c, v, False, _BLACK,     _WHITE,      'left',   'top',    _B_THIN)

# ─── Build Excel ──────────────────────────────────────────────────────────────

def build_excel(records, max_fields, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Tables'

    _main_hdr(ws.cell(1, 1), COL_FOLDER)
    _main_hdr(ws.cell(1, 2), COL_FILE)
    _main_hdr(ws.cell(1, 3), COL_HEADER)
    if max_fields > 1:
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + max_fields)
    ws.row_dimensions[1].height = 28

    excel_row = 2
    for rec in records:
        pairs = rec['_pairs']
        pad   = range(3 + len(pairs), 3 + max_fields)

        for ci, (lbl, _) in enumerate(pairs, 3): _sub_hdr(ws.cell(excel_row, ci), lbl)
        for ci in pad:                            _sub_hdr(ws.cell(excel_row, ci), '')
        _ff(ws.cell(excel_row, 1), rec['_folder'], True)
        _ff(ws.cell(excel_row, 2), rec['_file'],   True)
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        for ci, (_, val) in enumerate(pairs, 3): _data(ws.cell(excel_row, ci), val)
        for ci in pad:                            _data(ws.cell(excel_row, ci), '')
        _ff(ws.cell(excel_row, 1), None, False)
        _ff(ws.cell(excel_row, 2), None, False)
        ws.row_dimensions[excel_row].height = 18

        ws.merge_cells(start_row=excel_row-1, start_column=1, end_row=excel_row, end_column=1)
        ws.merge_cells(start_row=excel_row-1, start_column=2, end_row=excel_row, end_column=2)
        excel_row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 42
    for i in range(max_fields):
        ws.column_dimensions[get_column_letter(3 + i)].width = 32

    wb.save(output_path)
    print(f'\nSaved: {output_path}  ({len(records)} records, {max_fields} fields)')
    for r in records:
        tag = ' | '.join(f'{l}={v!r}' for l, v in r['_pairs']) or '(no header)'
        print(f'  {r["_file"]}: {tag}')

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    INPUT_FOLDER = Path(__file__).parent / 'input'
    OUTPUT_PATH  = Path(__file__).parent / 'output' / 'header_extract_formatted.xlsx'

    if not INPUT_FOLDER.exists():
        sys.exit(f'[ERROR] Input folder not found: {INPUT_FOLDER}')

    print(f'Scanning: {INPUT_FOLDER}')
    records, max_fields = scan_word_folder(INPUT_FOLDER)

    if not records:
        print('No .docx or .doc files found.')
    else:
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        build_excel(records, max_fields, OUTPUT_PATH)