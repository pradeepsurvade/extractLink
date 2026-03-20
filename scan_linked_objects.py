"""
Linked Object Scanner — SharePoint Edition
===========================================
Connects to a SharePoint document library via Microsoft Graph API,
downloads all .docx files (recursively through subfolders), scans each
for embedded/linked OLE objects, extracts them locally, and writes a report.

Prerequisites (install once):
    pip install msal requests lxml openpyxl

Authentication:
    Uses interactive browser SSO (device-code flow). On first run a browser
    window opens — log in with your Microsoft/SSO account. The token is
    cached in token_cache.json so subsequent runs skip the login prompt.

Configuration:
    Edit the CONFIG block below, then run:
        python scan_linked_objects.py
"""

import os, io, re, json, shutil, zipfile, urllib.parse, tempfile
from datetime import datetime
from pathlib import Path

import msal, requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── CONFIG ────────────────────────────────────────────────────────────────────
SHAREPOINT_URL  = "https://yourcompany.sharepoint.com/sites/YourSite"
SHAREPOINT_PATH = "/Shared Documents/YourFolder"   # library path to scan
OUTPUT_FOLDER   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Output")
REPORT_NAME     = "embedded_objects_report.xlsx"
TOKEN_CACHE     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "token_cache.json")

# Azure app registration — use the public Microsoft Office client ID
# (works for any user with a Microsoft/O365 account, no app registration needed)
CLIENT_ID = "d3590ed6-52b3-4102-aeff-aad2292ab01c"   # Microsoft Office public client
AUTHORITY = "https://login.microsoftonline.com/common"
SCOPES    = ["https://graph.microsoft.com/Files.Read.All",
             "https://graph.microsoft.com/Sites.Read.All"]
# ─────────────────────────────────────────────────────────────────────────────

NS_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_W   = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
NS_V   = 'urn:schemas-microsoft-com:vml'


# ── SharePoint / Graph authentication ─────────────────────────────────────────

def get_token():
    """
    Acquire an access token via interactive browser SSO (device-code flow).
    Token is cached in token_cache.json — subsequent runs reuse it silently.
    """
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE):
        cache.deserialize(open(TOKEN_CACHE).read())

    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)

    # Try silent (cached) token first
    accounts = app.get_accounts()
    result   = app.acquire_token_silent(SCOPES, account=accounts[0]) if accounts else None

    if not result:
        # Interactive device-code login
        flow   = app.initiate_device_flow(scopes=SCOPES)
        print(f"\n{flow['message']}\n")          # prints URL + code to enter
        result = app.acquire_token_by_device_flow(flow)

    if "access_token" not in result:
        raise RuntimeError(f"Authentication failed: {result.get('error_description')}")

    # Persist cache for next run
    open(TOKEN_CACHE, "w").write(cache.serialize())
    return result["access_token"]


def graph_get(token, url, **params):
    """Make an authenticated GET request to Microsoft Graph API."""
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params=params or None,
        timeout=30
    )
    resp.raise_for_status()
    return resp.json()


def graph_download(token, url):
    """Download file bytes from a Graph download URL."""
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    resp.raise_for_status()
    return resp.content


# ── Enumerate .docx files in SharePoint ───────────────────────────────────────

def get_sharepoint_docx_files(token):
    """
    Resolve the SharePoint site + library path to a Graph drive item,
    then recursively list all .docx files, returning:
        [{'name': ..., 'download_url': ..., 'relative_path': ...}, ...]
    """
    # Parse site hostname and path from SHAREPOINT_URL
    parsed   = urllib.parse.urlparse(SHAREPOINT_URL)
    hostname = parsed.netloc
    site_path = parsed.path.rstrip('/')

    # Get Graph site ID
    site     = graph_get(token, f"https://graph.microsoft.com/v1.0/sites/{hostname}:{site_path}")
    site_id  = site["id"]

    # Get default drive
    drives   = graph_get(token, f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives")
    drive_id = drives["value"][0]["id"]

    # Resolve the folder path inside the drive
    folder_path = SHAREPOINT_PATH.strip('/')
    folder_url  = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_path}"
    folder      = graph_get(token, folder_url)
    folder_id   = folder["id"]

    # Recursively collect .docx files
    return _list_docx_recursive(token, drive_id, folder_id, "")


def _list_docx_recursive(token, drive_id, folder_id, rel_path):
    """Recursively walk drive folders and return all .docx file metadata."""
    items_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_id}/children"
    data      = graph_get(token, items_url)
    files     = []

    for item in data.get("value", []):
        name     = item["name"]
        item_rel = f"{rel_path}/{name}".lstrip('/')

        if "folder" in item:
            # Recurse into subfolder
            files.extend(_list_docx_recursive(token, drive_id, item["id"], item_rel))
        elif name.lower().endswith('.docx') and not name.startswith('~'):
            files.append({
                "name":          name,
                "relative_path": item_rel,
                "download_url":  item.get("@microsoft.graph.downloadUrl") or
                                 f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item['id']}/content"
            })

    return files


# ── Utilities ─────────────────────────────────────────────────────────────────

def ole_type(prog, ext):
    if 'Excel' in prog or ext in ('.xlsx', '.xls', '.xlsm'): return 'Excel'
    if 'Acro'  in prog or 'PDF' in prog or ext == '.pdf':    return 'PDF'
    if 'Word'  in prog or ext in ('.docx', '.doc'):          return 'Word'
    return ext.upper().lstrip('.') or 'Object'

def emf_display_name(data):
    decoded = [s.decode('utf-16-le') for s in re.findall(b'(?:[\x20-\x7e]\x00){3,}', data)]
    found   = []
    for i, s in enumerate(decoded):
        if s == 'IconOnly':
            for c in decoded[max(0, i-6):i]:
                if '.' in c and '\\' not in c and c != 'IconOnly':
                    found.append(c)
    return max(found, key=len) if found else ''

def unique_dest(folder, name):
    dest = os.path.join(folder, name)
    if os.path.exists(dest):
        b, e = os.path.splitext(name)
        dest = os.path.join(folder, f"{b}_{datetime.now().strftime('%H%M%S%f')}{e}")
    return dest

def short_path(p):
    n = p.replace('\\', '/')
    i = n.find(f'/{os.path.basename(OUTPUT_FOLDER)}/')
    return '..' + n[i:] if i >= 0 else p

def make_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Parse .docx (from bytes) ──────────────────────────────────────────────────

def parse_docx(docx_bytes, docx_name):
    """Same as before but accepts raw bytes instead of a file path."""
    with zipfile.ZipFile(io.BytesIO(docx_bytes)) as z:
        doc_xml, rels_xml, names = (z.read("word/document.xml"),
                                    z.read("word/_rels/document.xml.rels"),
                                    z.namelist())
        emf = {os.path.basename(n): z.read(n) for n in names
               if n.startswith('word/media/') and n.endswith('.emf')}

    rels = {r.get('Id'): {'type': r.get('Type','').split('/')[-1], 'target': r.get('Target','')}
            for r in etree.fromstring(rels_xml).findall(f'{{{NS_PKG}}}Relationship')}
    hlinks = {rid: v['target'] for rid, v in rels.items() if v['type'] == 'hyperlink'}

    tree = etree.fromstring(doc_xml)
    body = tree.find(f'{{{NS_W}}}body')
    body_kids = list(body) if body is not None else []

    shape_names = {}
    for shape in tree.iter(f'{{{NS_V}}}shape'):
        img = shape.find(f'{{{NS_V}}}imagedata')
        if img is None: continue
        tgt  = rels.get(img.get(f'{{{NS_R}}}id',''), {}).get('target','')
        name = emf_display_name(emf.get(os.path.basename(tgt), b''))
        if name: shape_names[shape.get('id','')] = name

    def text(e): return ''.join(t.text or '' for t in e.iter(f'{{{NS_W}}}t')).strip()

    def section(ole):
        idx = next((i for i, c in enumerate(body_kids) if any(x is ole for x in c.iter())), None)
        if idx is None: return "DOCUMENT"
        for c in reversed(body_kids[:idx]):
            if c.tag == f'{{{NS_W}}}p':
                ps = c.find(f'.//{{{NS_W}}}pStyle')
                if ps is not None and 'eading' in ps.get(f'{{{NS_W}}}val',''):
                    t = text(c)
                    if t: return t.upper()
        return "DOCUMENT"

    def row_label(ole):
        p = ole.getparent()
        while p is not None:
            if p.tag.split('}')[-1] == 'tr':
                return next((text(tc) for tc in p.findall(f'.//{{{NS_W}}}tc') if text(tc)), '')[:100]
            p = p.getparent()
        return ''

    def hyperlink(ole):
        p = ole.getparent()
        while p is not None:
            if p.tag.split('}')[-1] == 'hyperlink':
                return hlinks.get(p.get(f'{{{NS_R}}}id',''),'')
            p = p.getparent()
        return ''

    def internal_path(rid, obj_type, ext):
        c = 'word/' + rels[rid]['target'].lstrip('/')
        if c in names: return c
        if obj_type == 'PDF':
            b = next((n for n in names if 'embeddings' in n and n.endswith('.bin')), None)
            if b: return b
        return next((n for n in names if 'embeddings' in n and n.endswith(ext)), None)

    results = []
    for ole in tree.iter():
        if ole.tag.split('}')[-1] != 'OLEObject': continue
        rid  = ole.get(f'{{{NS_R}}}id','')
        prog = ole.get('ProgID','')
        if not rid or rid not in rels: continue

        raw     = rels[rid]['target']
        decoded = urllib.parse.unquote(raw)
        is_url  = decoded.startswith(('http://','https://'))
        if decoded.startswith('file:///'): decoded = decoded[8:]

        ext      = os.path.splitext(os.path.basename(decoded))[1].lower()
        obj_type = ole_type(prog, ext)
        url      = hyperlink(ole) or (urllib.parse.unquote(raw) if is_url else '')

        results.append({
            'source_docx':   docx_name,
            'object_file':   shape_names.get(ole.get('ShapeID',''),'') or os.path.basename(decoded),
            'object_type':   obj_type,
            'section':       section(ole),
            'row_label':     row_label(ole),
            'hyperlink_url': url,
            'internal_path': None if is_url else internal_path(rid, obj_type, ext),
            'ole_target':    urllib.parse.unquote(raw),
            'docx_bytes':    docx_bytes,  # kept for extraction step
        })
    return results


# ── Fix & verify extracted Excel files ────────────────────────────────────────

def fix_excel(path):
    """Strip Word OLE embedding flags so the file opens as a normal workbook."""
    try:
        buf = io.BytesIO(open(path,'rb').read())
        with zipfile.ZipFile(buf) as z:
            if 'xl/workbook.xml' not in z.namelist(): return
            xml = z.read('xl/workbook.xml')
        if b'visibility="hidden"' not in xml and b'<oleSize' not in xml and b'activeTab=' in xml: return
        xml = xml.replace(b' visibility="hidden"',b'').replace(b'visibility="hidden"',b'')
        if b'activeTab=' not in xml: xml = xml.replace(b'<workbookView',b'<workbookView activeTab="0"')
        while b'<oleSize' in xml:
            s, e = xml.find(b'<oleSize'), xml.find(b'>',xml.find(b'<oleSize'))+1
            xml  = xml[:s]+xml[e:]
        out = io.BytesIO()
        buf.seek(0)
        with zipfile.ZipFile(buf) as zi, zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED) as zo:
            for item in zi.infolist():
                zo.writestr(item, xml if item.filename=='xl/workbook.xml' else zi.read(item.filename))
        open(path,'wb').write(out.getvalue())
    except Exception: pass

def verify(path):
    if not os.path.isfile(path) or os.path.getsize(path) == 0:
        return False, "missing or empty"
    size, ext = os.path.getsize(path), os.path.splitext(path)[1].lower()
    if ext in ('.xlsx','.xlsm','.xls'):
        try:
            with zipfile.ZipFile(path) as z:
                bad = z.testzip()
                if bad: return False, f"corrupt: {bad}"
                if 'xl/workbook.xml' not in z.namelist(): return False, "no workbook.xml"
                if b'visibility="hidden"' in z.read('xl/workbook.xml'): return False, "still hidden"
                cells = sum(len(re.findall(b'<c ',z.read(n))) for n in z.namelist() if 'xl/worksheets/sheet' in n)
            return (True, f"OK ({size:,}b, {cells} cells)") if cells else (False,"no data")
        except zipfile.BadZipFile: return False, "invalid xlsx"
    if ext == '.pdf':
        with open(path,'rb') as f: hdr = f.read(4)
        return (True, f"OK ({size:,}b)") if hdr == b'%PDF' else (False,"invalid PDF")
    return True, f"OK ({size:,}b)"


# ── Extract embedded files ─────────────────────────────────────────────────────

def extract_files(objects):
    """Extract embedded OLE files from the in-memory docx bytes to Output/<DocName>/."""
    for obj in objects:
        if obj['hyperlink_url']:
            obj['saved_path'] = ''; continue

        docx_bytes = obj.pop('docx_bytes')   # consume — no longer needed after extraction
        stem       = os.path.splitext(obj['source_docx'])[0]
        out_dir    = os.path.join(OUTPUT_FOLDER, stem)
        os.makedirs(out_dir, exist_ok=True)

        if not obj['internal_path']:
            obj['saved_path'] = ''; obj['verify_status'] = ''; continue

        name = obj['object_file']
        if obj['internal_path'].endswith('.bin') and name.lower().endswith('.bin'):
            type_ext = '.pdf' if obj['object_type'] == 'PDF' else '.bin'
            name     = f"{stem}_embedded{type_ext}"

        dest = unique_dest(out_dir, name)
        try:
            with zipfile.ZipFile(io.BytesIO(docx_bytes)) as z:
                data = z.read(obj['internal_path'])
            if obj['internal_path'].endswith('.bin') and obj['object_type'] == 'PDF':
                s, e = data.find(b'%PDF'), data.rfind(b'%%EOF')
                if s >= 0 and e >= 0: data = data[s:e+5]
            open(dest,'wb').write(data)
            if dest.lower().endswith(('.xlsx','.xlsm','.xls')): fix_excel(dest)
            obj['object_file']   = name
            obj['saved_path']    = os.path.abspath(dest)
            ok, msg              = verify(obj['saved_path'])
            obj['verify_status'] = f'✔ {msg}' if ok else f'✘ {msg}'
        except Exception as ex:
            obj['saved_path'] = ''; obj['verify_status'] = f'✘ {ex}'


# ── Write report ──────────────────────────────────────────────────────────────

def write_report(objects):
    wb = Workbook(); ws = wb.active; ws.title = "Embedded Objects"
    border = make_border()
    cols   = [("Word Document",35),("SharePoint Path",45),("Word Section",20),
              ("Row Label",55),("Object Type",14),("URL",70)]

    for ci,(h,w) in enumerate(cols,1):
        c = ws.cell(row=1,column=ci,value=h)
        c.font=Font(name="Arial",size=11,bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",start_color="808080")
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=border
        ws.column_dimensions[c.column_letter].width=w
    ws.row_dimensions[1].height=25

    def cell(r,c,v,color="000000",wrap=False):
        x=ws.cell(row=r,column=c,value=v)
        x.font=Font(name="Arial",size=10,color=color)
        x.fill=PatternFill("solid",start_color="FFFFFF")
        x.alignment=Alignment(horizontal="left",vertical="top",wrap_text=wrap)
        x.border=border

    for i,obj in enumerate(objects):
        r=i+2
        cell(r,1,obj['source_docx'])
        cell(r,2,obj.get('sharepoint_path',''))
        cell(r,3,obj['section'])
        cell(r,4,obj['row_label'],wrap=True)
        cell(r,5,obj['object_type'])
        url = obj['hyperlink_url'] or (short_path(obj['saved_path']) if obj.get('saved_path') else obj.get('ole_target',''))
        cell(r,6,url,color="0563C1",wrap=True)
        ws.row_dimensions[r].height=40 if len(obj.get('row_label',''))>40 else 18

    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:F{len(objects)+1}"
    path=os.path.join(OUTPUT_FOLDER,REPORT_NAME)
    wb.save(path); return path


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Step 1 — Authenticate
    print("\nAuthenticating with SharePoint...")
    token = get_token()
    print("✔ Authenticated")

    # Step 2 — List all .docx files in SharePoint
    print(f"\nScanning SharePoint: {SHAREPOINT_URL}{SHAREPOINT_PATH}")
    docx_files = get_sharepoint_docx_files(token)
    if not docx_files:
        print("No .docx files found."); return
    print(f"Found {len(docx_files)} .docx file(s)")

    all_objects = []

    # Step 3 — Download, parse, extract each file
    for i, item in enumerate(docx_files, 1):
        name = item['name']
        print(f"\n[{i}/{len(docx_files)}] {item['relative_path']}")

        try:
            docx_bytes = graph_download(token, item['download_url'])
            objects    = parse_docx(docx_bytes, name)
            # Tag each object with its SharePoint path for the report
            for obj in objects:
                obj['sharepoint_path'] = item['relative_path']
        except Exception as ex:
            print(f"  [!] Failed to download/parse: {ex}"); continue

        extract_files(objects)

        for obj in objects:
            if not obj['hyperlink_url']:
                print(f"  • {obj['object_file']}  [{obj.get('verify_status','')}]")

        all_objects.extend(objects)

    # Step 4 — Write report
    if all_objects:
        print(f"\nReport → {write_report(all_objects)}")

    total    = len(all_objects)
    with_url = sum(1 for o in all_objects if o['hyperlink_url'])
    print(f"\nScanned: {len(docx_files)} file(s) | Objects: {total} | Linked: {with_url} | Embedded: {total-with_url}\n")


if __name__ == "__main__":
    main()
