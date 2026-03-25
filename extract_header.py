import os, zipfile, re
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

W_NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

# ─── DOCX header extraction ───────────────────────────────────────────────────

def para_text(p):
    return re.sub(r'  +', ' ', ''.join(t.text for t in p.iter(f'{{{W_NS}}}t') if t.text)).strip()

def extract_header_fields(docx_path):
    """
    Dynamically extract ordered (label, value) pairs from the docx page header table.
    - Scans ALL headerN.xml files; picks the first with a table that has actual text.
    - Labels and values are read verbatim — nothing is hardcoded.
    - Cells with no label (no colon) but with a value are included as ('', value).

    Cell layouts handled:
      No colon, has text  : value-only cell          → ('', value)
      Single paragraph    : "Label: value"            → ('Label', 'value')
      Multi-paragraph     : ["Label:", "value"]       → ('Label:', 'value')
      Multi-paragraph     : ["Label: x", "value"]     → ('Label: x', 'value')
    """
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            header_files = sorted(
                n for n in z.namelist()
                if re.match(r'word/header\d+\.xml$', n)
            )
            header_xml = None
            for hf in header_files:
                content = z.read(hf)
                if b'<w:tbl' not in content:
                    continue
                root = ET.fromstring(content)
                texts = [t.text for t in root.iter(f'{{{W_NS}}}t') if t.text and t.text.strip()]
                if texts:
                    header_xml = content
                    break
            if header_xml is None:
                return []
        root = ET.fromstring(header_xml)
    except Exception:
        return []

    pairs = []
    for cell in root.iter(f'{{{W_NS}}}tc'):
        non_empty = [para_text(p) for p in cell.iter(f'{{{W_NS}}}p')]
        non_empty = [p for p in non_empty if p]

        if not non_empty:
            continue  # completely empty cell — skip

        colon_paras = [p for p in non_empty if ':' in p]

        if not colon_paras:
            # No label found — show value as-is with blank label
            value = ' '.join(non_empty).strip()
            pairs.append(('', value))
            continue

        lp = colon_paras[0]  # paragraph carrying the colon

        if len(non_empty) == 1:
            # Single paragraph: "Label: value" — split on first colon
            idx = lp.index(':')
            label = lp[:idx].strip()
            value = lp[idx + 1:].strip()
        else:
            # Multiple paragraphs: colon paragraph = label (verbatim), rest = value
            label = lp
            remaining = [p for p in non_empty if p != lp]
            value = ' '.join(remaining).strip()

        pairs.append((label, value))

    return pairs

def scan_docx_folder(folder):
    """Scan folder/subfolders for .docx files; return records and max_fields."""
    records = []
    max_fields = 0

    for docx_path in sorted(Path(folder).rglob('*.docx')):
        pairs = extract_header_fields(str(docx_path))
        max_fields = max(max_fields, len(pairs))
        records.append({
            '_folder': docx_path.parent.name,
            '_file':   docx_path.name,
            '_pairs':  pairs,
        })

    return records, max_fields

# ─── Excel formatting helpers ─────────────────────────────────────────────────

DARK_BLUE   = 'FF1F4E79'
MED_BLUE    = 'FF2E75B6'
LIGHT_GREY  = 'FFF2F2F2'
WHITE_FONT  = 'FFFFFFFF'
DARK_FONT   = 'FF1F4E79'
BORDER_DARK = 'FF4472C4'
BORDER_GREY = 'FFBFBFBF'

def thin_border(color):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border(color):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border_no_top(color):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, bottom=s)

def apply_main_header(cell, value):
    cell.value = value
    cell.font = Font(bold=True, color=WHITE_FONT, name='Calibri', size=11)
    cell.fill = PatternFill('solid', fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border(BORDER_GREY)

def apply_sub_header(cell, value):
    cell.value = value
    cell.font = Font(bold=True, color=WHITE_FONT, name='Calibri', size=11)
    cell.fill = PatternFill('solid', fgColor=MED_BLUE)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = thin_border(BORDER_GREY)

def apply_folder_file(cell, value, is_top_row):
    cell.value = value
    cell.font = Font(bold=True, color=DARK_FONT, name='Calibri', size=11)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREY)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = medium_border(BORDER_DARK) if is_top_row else medium_border_no_top(BORDER_DARK)

def apply_data(cell, value):
    cell.value = value
    cell.font = Font(name='Calibri', size=11)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = thin_border(BORDER_GREY)

# ─── Build Excel ──────────────────────────────────────────────────────────────

def build_excel(records, max_fields, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Tables'

    # Row 1: Folder Name | Word Doc Name | Header Content (merged across max_fields cols)
    apply_main_header(ws.cell(1, 1), 'Folder Name')
    apply_main_header(ws.cell(1, 2), 'Word Doc Name')
    apply_main_header(ws.cell(1, 3), 'Header Content')
    if max_fields > 1:
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + max_fields)
    ws.row_dimensions[1].height = 28

    # 2 rows per record:
    #   Row A — Folder | File | label1 | label2 | ...  (this file's own labels)
    #   Row B — merged | merged | value1 | value2 | ... (this file's own values)
    excel_row = 2
    for rec in records:
        pairs = rec['_pairs']

        # Sub-header: this file's own labels verbatim (blank string if no label in doc)
        for ci, (label, _) in enumerate(pairs, start=3):
            apply_sub_header(ws.cell(excel_row, ci), label)
        for ci in range(3 + len(pairs), 3 + max_fields):
            apply_sub_header(ws.cell(excel_row, ci), '')
        apply_folder_file(ws.cell(excel_row, 1), rec['_folder'], is_top_row=True)
        apply_folder_file(ws.cell(excel_row, 2), rec['_file'],   is_top_row=True)
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        # Values: this file's own values (always shown, even when label is blank)
        for ci, (_, value) in enumerate(pairs, start=3):
            apply_data(ws.cell(excel_row, ci), value)
        for ci in range(3 + len(pairs), 3 + max_fields):
            apply_data(ws.cell(excel_row, ci), '')
        apply_folder_file(ws.cell(excel_row, 1), None, is_top_row=False)
        apply_folder_file(ws.cell(excel_row, 2), None, is_top_row=False)
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        # Merge Folder Name and File Name across both rows
        ws.merge_cells(start_row=excel_row-2, start_column=1, end_row=excel_row-1, end_column=1)
        ws.merge_cells(start_row=excel_row-2, start_column=2, end_row=excel_row-1, end_column=2)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 42
    for i in range(max_fields):
        ws.column_dimensions[get_column_letter(3 + i)].width = 32

    wb.save(output_path)
    print(f'Saved: {output_path}')
    print(f'  Records   : {len(records)}')
    print(f'  Max fields: {max_fields}')
    for r in records:
        print(f'  {r["_file"]}:')
        for label, value in r['_pairs']:
            print(f'    label={label!r:45s} → value={value!r}')

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    INPUT_FOLDER = '/mnt/user-data/uploads'
    OUTPUT_PATH  = '/mnt/user-data/outputs/header_extract_formatted.xlsx'

    records, max_fields = scan_docx_folder(INPUT_FOLDER)
    if not records:
        print('No .docx files found.')
    else:
        os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
        build_excel(records, max_fields, OUTPUT_PATH)
