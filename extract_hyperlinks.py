"""
Hyperlink Extractor for Word Documents (.docx / .doc)
Single sheet 'Embedded Objects' — visual grouping via borders only (filterable).
URLs: http/https/mailto/ftp only. Links not starting with http are excluded.

.docx -> XML extraction (section, table, row, column context)
.doc  -> Pure-Python OLE CFB reader (stdlib struct only, zero pip installs)
"""

from pathlib import Path
from itertools import groupby
import zipfile, struct, re

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
INPUT_FOLDER = Path(__file__).parent / 'input'
OUTPUT_PATH  = Path(__file__).parent / 'output' / 'Hyperlink_extract_formatted.xlsx'

# ── XML namespaces ─────────────────────────────────────────────────────────
NS = {
    'w':   'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}
HEADING_STYLES = {f'heading{i}' for i in range(1, 7)}
SECTION_STYLES = HEADING_STYLES | {'title', 'subtitle'}

# ── OLE CFB constants ──────────────────────────────────────────────────────
OLE_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
FREESECT  = 0xFFFFFFFF
ENDOFCHAIN= 0xFFFFFFFE
DIFSECT   = 0xFFFFFFFC
FATSECT   = 0xFFFFFFFD

# Only keep URLs that start with http (http/https)
URL_PAT = re.compile(
    r'(https?://[^\x00-\x08\x0a-\x1f]{5,})'
)

# ══════════════════════════════════════════════════════════════════════════
#  Pure-Python .doc parser — OLE reader + Word Binary Format text extractor
#  stdlib only (struct, re). No pip install, no LibreOffice needed.
#  Handles ASCII/cp1252 and UTF-16LE encoded .doc files.
#  Extracts section / table / row / column context from cell markers and
#  field codes embedded in the WordDocument stream.
# ══════════════════════════════════════════════════════════════════════════

FREESECT   = 0xFFFFFFFF
ENDOFCHAIN = 0xFFFFFFFE
DIFSECT    = 0xFFFFFFFC
FATSECT    = 0xFFFFFFFD
OLE_MAGIC  = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'

CELL_MARK = '\x07'
PARA_MARK = '\r'
FLD_BEGIN  = '\x13'
FLD_SEP    = '\x14'
FLD_END    = '\x15'

# Section heading: ALL-CAPS words ≥ 5 chars, optionally prefixed by a number
HEADING_RE  = re.compile(r'^(\d+[\.\s]+)?[A-Z][A-Z\s/\-\u2013\u2014&()]{4,}$')
NOT_HEADING = re.compile(
    r'^(Note|Reference|Refer|Include|List|Embed|Provide|NA\b|Yes\b|No\b)',
    re.IGNORECASE)

SKIP_PREFIXES = (
    'http://schemas.', 'http://www.w3.', 'http://purl.',
    'https://schemas.', 'http://schemas.microsoft',
    'http://schemas.openxmlformats', 'http://dublincore.org/schemas',
)


# ══════════════════════════════════════════════════════════════════════════
#  OLE reader
# ══════════════════════════════════════════════════════════════════════════

def _read_ole_streams_full(path):
    with open(path, 'rb') as fh:
        data = fh.read()
    if data[:8] != OLE_MAGIC:
        return {}

    sec_size      = 1 << struct.unpack_from('<H', data, 30)[0]
    mini_ss       = 1 << struct.unpack_from('<H', data, 32)[0]
    dir_start     = struct.unpack_from('<I', data, 48)[0]
    mini_cutoff   = struct.unpack_from('<I', data, 56)[0]
    minifat_start = struct.unpack_from('<I', data, 60)[0]
    header_difat  = list(struct.unpack_from('<109I', data, 76))
    fat_sids      = [s for s in header_difat
                     if s not in (FREESECT, ENDOFCHAIN, DIFSECT, FATSECT)]

    difat_start = struct.unpack_from('<I', data, 68)[0]
    if difat_start not in (FREESECT, ENDOFCHAIN):
        sec = difat_start
        while sec not in (FREESECT, ENDOFCHAIN):
            off     = 512 + sec * sec_size
            entries = list(struct.unpack_from('<' + 'I' * (sec_size // 4), data, off))
            fat_sids.extend(s for s in entries[:-1]
                            if s not in (FREESECT, ENDOFCHAIN, DIFSECT, FATSECT))
            sec = entries[-1]

    def sec_off(sid): return 512 + sid * sec_size

    fat = []
    for sid in fat_sids:
        fat.extend(struct.unpack_from('<' + 'I' * (sec_size // 4), data, sec_off(sid)))

    def read_chain(start):
        out, sec, seen = [], start, set()
        while sec not in (ENDOFCHAIN, FREESECT) and sec < len(fat):
            if sec in seen: break
            seen.add(sec)
            out.append(data[sec_off(sec):sec_off(sec) + sec_size])
            sec = fat[sec]
        return b''.join(out)

    dir_data = read_chain(dir_start)
    minifat  = []
    if minifat_start not in (ENDOFCHAIN, FREESECT):
        mf      = read_chain(minifat_start)
        minifat = list(struct.unpack_from('<' + 'I' * (len(mf) // 4), mf))

    mini_stream = read_chain(struct.unpack_from('<I', dir_data, 116)[0])

    def read_mini(start, sz):
        out, sec, seen = [], start, set()
        while sec not in (ENDOFCHAIN, FREESECT) and sec < len(minifat):
            if sec in seen: break
            seen.add(sec)
            o = sec * mini_ss
            out.append(mini_stream[o:o + mini_ss])
            sec = minifat[sec]
        return b''.join(out)[:sz]

    streams = {}
    for i in range(len(dir_data) // 128):
        e   = dir_data[i * 128:(i + 1) * 128]
        nl  = struct.unpack_from('<H', e, 64)[0]
        typ = e[66]
        ss  = struct.unpack_from('<I', e, 116)[0]
        sz  = struct.unpack_from('<I', e, 120)[0]
        if typ == 2 and nl > 0 and ss not in (ENDOFCHAIN, FREESECT):
            name    = e[:nl - 2].decode('utf-16-le', errors='replace')
            content = (read_mini(ss, sz) if sz < mini_cutoff
                       else read_chain(ss))[:sz]
            streams[name.lower()] = content

    return streams


def _decode_worddocument(streams):
    """Auto-detect encoding (ASCII/cp1252 vs UTF-16LE) and decode."""
    wd = streams.get('worddocument', b'')
    if not wd:
        return '', 'unknown'
    if b'https://' in wd or b'http://' in wd:
        return wd.decode('cp1252', errors='replace'), 'cp1252'
    http16  = 'http://'.encode('utf-16-le')
    https16 = 'https://'.encode('utf-16-le')
    if http16 in wd or https16 in wd:
        return wd.decode('utf-16-le', errors='replace'), 'utf-16le'
    try:
        text = wd.decode('utf-16-le', errors='replace')
        if any(c.isalpha() for c in text[:500]):
            return text, 'utf-16le'
    except Exception:
        pass
    return wd.decode('cp1252', errors='replace'), 'cp1252'


# ══════════════════════════════════════════════════════════════════════════
#  Text helpers
# ══════════════════════════════════════════════════════════════════════════

def _get_urls(txt):
    """Extract http/https URLs from field codes and bare text in a cell."""
    urls, seen = [], set()

    def add(url):
        url = url.strip().rstrip('"').rstrip("'")
        if (url.startswith('http')
                and not any(url.startswith(p) for p in SKIP_PREFIXES)
                and url not in seen):
            seen.add(url)
            urls.append(url)

    # Field-code hyperlinks: \x13 HYPERLINK "url" \x14 display \x15
    for m in re.finditer(
            re.escape(FLD_BEGIN) + r'(.*?)' + re.escape(FLD_END),
            txt, re.DOTALL):
        field   = m.group(1)
        sep_idx = field.find(FLD_SEP)
        instr   = field[:sep_idx] if sep_idx >= 0 else field
        um = re.search(
            r'HYPERLINK\s+"?(https?://[^"\s\x13\x14\x15\x07\r]+)',
            instr, re.IGNORECASE)
        if um:
            add(um.group(1))

    # Bare URL fallback (some docs omit field codes)
    for m in re.finditer(r'(?<!["\w])(https?://[^\s\x07\x13\x14\x15\r"]{8,})', txt):
        add(m.group(1))

    return urls


def _clean(txt):
    """Remove field codes and return printable text."""
    c = re.sub(re.escape(FLD_BEGIN) + r'.*?' + re.escape(FLD_END),
               '', txt, flags=re.DOTALL)
    return ''.join(ch for ch in c if ch.isprintable()).strip()


def _is_heading(text):
    t = text.strip()
    if len(t) < 5:
        return False
    if ' ' not in t and len(t) < 8:   # single short word — skip (e.g. 'ASKB')
        return False
    if NOT_HEADING.match(t):
        return False
    return bool(HEADING_RE.match(t))


# ══════════════════════════════════════════════════════════════════════════
#  Table parser
# ══════════════════════════════════════════════════════════════════════════

def _parse_table(row_texts, section):
    """
    Parse table row strings into hyperlink records.

    In Word binary format:
    - Cells are separated by \x07
    - A double \x07\x07 separates the header row from data rows when both
      appear in the same paragraph (common in Word 97-2003 .doc files)
    - Row labels come from the first non-empty cell of each row
    - Column labels come from the header row (first row with no URLs)
    """
    records = []

    # Expand paragraphs that contain multiple logical rows (split on \x07\x07)
    expanded = []
    for rt in row_texts:
        parts = rt.split(CELL_MARK + CELL_MARK)
        for p in parts:
            if p:
                expanded.append(p)

    if not expanded:
        return records

    # Detect header row
    header_labels = []
    data_start    = 0
    if expanded and not _get_urls(expanded[0]):
        raw_hdr   = [c for c in expanded[0].split(CELL_MARK)]
        hdr_clean = [_clean(c) for c in raw_hdr]
        if any(h for h in hdr_clean):
            header_labels = hdr_clean
            data_start    = 1

    # Column count from header
    n_cols = len([h for h in header_labels if h]) if header_labels else 0

    for row_text in expanded[data_start:]:
        cells = row_text.split(CELL_MARK)

        # Use first non-empty cell as row_label (handles rows starting with empty cell)
        row_label = ''
        for c in cells:
            lbl = _clean(c)
            if lbl:
                row_label = lbl
                break

        # If many cells and we know col count, iterate in strides
        stride = n_cols if n_cols >= 2 else len(cells)
        for start_ci in range(0, len(cells), max(stride, 1)):
            chunk = cells[start_ci:start_ci + stride]
            # row_label for sub-chunk = first non-empty clean cell
            rl = next((_clean(c) for c in chunk if _clean(c)), row_label)

            for ci, cell in enumerate(chunk):
                abs_ci = start_ci + ci
                for url in _get_urls(cell):
                    rl_col  = header_labels[0]      if header_labels else ''
                    url_col = (header_labels[abs_ci]
                               if abs_ci < len(header_labels) else '')
                    records.append({
                        'section':       section,
                        'table_flag':    'Within Table',
                        'row_label':     rl,
                        'row_label_col': rl_col,
                        'url_col':       url_col,
                        'url':           url,
                    })

    return records


# ══════════════════════════════════════════════════════════════════════════
#  Main extractor
# ══════════════════════════════════════════════════════════════════════════

def extract_from_doc(path):
    """
    Extract hyperlinks from a .doc file using pure Python (no pip required).
    Returns list of record dicts with keys:
      section, table_flag, row_label, row_label_col, url_col, url
    """
    streams = _read_ole_streams_full(path)
    text, _ = _decode_worddocument(streams)
    if not text:
        return []

    raw_paras = text.split(PARA_MARK)

    # ── Build logical units ───────────────────────────────────────────────
    # Rules:
    # 1. Para with \x07 = table content; accumulate until row ends with \x07
    # 2. Para WITHOUT \x07 after a still-open table row = cell continuation
    # 3. Empty para breaks a cell continuation and closes the row
    # 4. Non-empty para without \x07 after a closed table = plain paragraph

    units    = []
    row_acc  = ''        # accumulates a multi-\r table row
    in_table = False     # True while inside a table block

    for para in raw_paras:
        if CELL_MARK in para:
            in_table = True
            row_acc  = (row_acc + para) if row_acc else para
            # Row complete when this para ends with \x07 (= row terminator before \r)
            if para.endswith(CELL_MARK):
                units.append({'type': 'table_row', 'text': row_acc})
                row_acc = ''
            # else: row continues — next para may be cell content or end-of-row

        elif in_table:
            # No cell marks in this para
            if row_acc:
                # We had an unfinished row; this para is continuation text
                row_acc += ' ' + para
                # Close the row now — next para will decide whether still in table
                units.append({'type': 'table_row', 'text': row_acc})
                row_acc = ''
                # Stay in_table=True; next non-empty non-cell para will end table
            else:
                # Between rows: empty para = still in table; non-empty = table ended
                if para.strip():
                    in_table = False
                    units.append({'type': 'para', 'text': para})
                # else: skip empty inter-row para silently

        else:
            units.append({'type': 'para', 'text': para})

    if row_acc:
        units.append({'type': 'table_row', 'text': row_acc})

    # ── Walk units ────────────────────────────────────────────────────────
    current_section = ''
    records         = []
    table_buffer    = []

    def flush():
        nonlocal table_buffer
        if table_buffer:
            records.extend(
                _parse_table([u['text'] for u in table_buffer], current_section))
            table_buffer = []

    for unit in units:
        if unit['type'] == 'table_row':
            table_buffer.append(unit)
        else:
            flush()
            txt   = unit['text']
            clean = _clean(txt)
            if _is_heading(clean):
                current_section = clean
            for url in _get_urls(txt):
                records.append({
                    'section':       current_section,
                    'table_flag':    'Outside Table',
                    'row_label':     '',
                    'row_label_col': '',
                    'url_col':       '',
                    'url':           url,
                })

    flush()
    return records


# ══════════════════════════════════════════════════════════════════════════
#  DOCX extractor
# ══════════════════════════════════════════════════════════════════════════

def _txt(el):
    return ' '.join(
        (t.text or '').strip()
        for t in el.findall('.//{%s}t' % NS['w'])
        if (t.text or '').strip()
    )

def _load_rels(z):
    rels = {}
    try:
        with z.open('word/_rels/document.xml.rels') as f:
            tree = etree.parse(f)
        for rel in tree.findall('{%s}Relationship' % NS['rel']):
            if 'hyperlink' in rel.get('Type', ''):
                rels[rel.get('Id')] = rel.get('Target', '')
    except KeyError:
        pass
    return rels


def _build_num_map(z, body):
    """
    Build a map {para_element_id -> number_string} for every paragraph
    that carries a w:numPr (auto-numbered by Word).

    Reads word/numbering.xml to get each list's start value, then walks
    all paragraphs in document order counting increments — exactly what
    Word renders on screen.  Only decimal / ordinal formats are supported
    (sufficient for section headings numbered 1, 2, 3…).

    Returns dict keyed by id(para_element).
    """
    # ── Read abstractNum start values from numbering.xml ──────────────
    abs_starts = {}   # abstractNumId -> {ilvl -> int}
    num_to_abs = {}   # numId         -> abstractNumId

    try:
        with z.open('word/numbering.xml') as f:
            num_tree = etree.parse(f)
    except KeyError:
        return {}

    for an in num_tree.findall('{%s}abstractNum' % NS['w']):
        anid = an.get('{%s}abstractNumId' % NS['w'])
        abs_starts[anid] = {}
        for lvl in an.findall('{%s}lvl' % NS['w']):
            ilvl     = lvl.get('{%s}ilvl' % NS['w'])
            start_el = lvl.find('{%s}start' % NS['w'])
            fmt_el   = lvl.find('{%s}numFmt' % NS['w'])
            fmt      = fmt_el.get('{%s}val' % NS['w']) if fmt_el is not None else 'decimal'
            start    = int(start_el.get('{%s}val' % NS['w'])) if start_el is not None else 1
            # Only track decimal/ordinal lists (section numbers)
            if fmt in ('decimal', 'decimalZero', 'ordinal', 'cardinalText'):
                abs_starts[anid][ilvl] = start

    for num in num_tree.findall('{%s}num' % NS['w']):
        nid   = num.get('{%s}numId' % NS['w'])
        anid_el = num.find('{%s}abstractNumId' % NS['w'])
        if anid_el is not None:
            anid = anid_el.get('{%s}val' % NS['w'])
            num_to_abs[nid] = anid
        # Apply startOverride if present
        for ov in num.findall('{%s}lvlOverride' % NS['w']):
            ilvl_ov = ov.get('{%s}ilvl' % NS['w'])
            so = ov.find('{%s}startOverride' % NS['w'])
            if so is not None:
                anid_v = num_to_abs.get(nid)
                if anid_v and anid_v in abs_starts:
                    abs_starts[anid_v][ilvl_ov] = int(so.get('{%s}val' % NS['w']))

    # ── Walk body direct children — same element objects as main extract loop ──
    # Using the element itself as dict key (lxml creates different Python wrapper
    # objects for the same XML node on each traversal, making id() unreliable).
    counters = {}   # (numId, ilvl) -> current int value
    num_map  = {}   # element -> number string

    for para in body:   # direct children only — mirrors main extract_from_docx loop
        if para.tag.split('}')[-1] != 'p':
            continue
        pPr   = para.find('{%s}pPr' % NS['w'])
        numPr = pPr.find('{%s}numPr' % NS['w']) if pPr is not None else None
        if numPr is None:
            continue
        nid_el  = numPr.find('{%s}numId' % NS['w'])
        ilvl_el = numPr.find('{%s}ilvl' % NS['w'])
        nid  = nid_el.get('{%s}val' % NS['w'])  if nid_el  is not None else None
        ilvl = ilvl_el.get('{%s}val' % NS['w']) if ilvl_el is not None else '0'
        if not nid or nid == '0':
            continue
        anid = num_to_abs.get(nid)
        if not anid or ilvl not in abs_starts.get(anid, {}):
            continue

        key = (nid, ilvl)
        counters[key] = counters[key] + 1 if key in counters else abs_starts[anid][ilvl]
        num_map[para] = str(counters[key])   # element as key — always same object

    return num_map

def _pstyle(para):
    ps = para.find('{%s}pPr/{%s}pStyle' % (NS['w'], NS['w']))
    return (ps.get('{%s}val' % NS['w']) or '').lower() if ps is not None else ''

def _is_heading_style(para):
    """True if paragraph has a Heading / Title / Subtitle style."""
    style = _pstyle(para)
    return style in SECTION_STYLES or 'heading' in style

def _is_header_row(row):
    """
    True if a table row is a true header row (column labels), not a data row.
    Checks three signals — any one is sufficient:
      1. w:tblHeader marker in trPr  (explicit Word repeat-header flag)
      2. ALL cells have a non-trivial background shading (header fill colour)
      3. ALL non-empty cells are bold (bold labels = header pattern)
    """
    trPr = row.find('{%s}trPr' % NS['w'])
    if trPr is not None and trPr.find('{%s}tblHeader' % NS['w']) is not None:
        return True
    cells = row.findall('{%s}tc' % NS['w'])
    if not cells:
        return False
    NON_HEADER_FILLS = {None, 'auto', 'ffffff', 'FFFFFF', '000000', '00000000'}
    shadings = [
        (cell.find('.//{%s}shd' % NS['w']) is not None and
         cell.find('.//{%s}shd' % NS['w']).get('{%s}fill' % NS['w']) not in NON_HEADER_FILLS)
        for cell in cells
    ]
    if all(shadings):
        return True
    non_empty = [(c, _txt(c)) for c in cells if _txt(c)]
    if non_empty and all(bool(c.findall('.//{%s}b' % NS['w'])) for c, _ in non_empty):
        return True
    return False


def _hl_urls(el, rels):
    """
    Return all http/https URLs from an element via two mechanisms:

    1. <w:hyperlink r:id="rIdN"> — standard relationship-based hyperlinks.
    2. Runs with <w:rStyle w:val="Hyperlink"> — plain runs styled as hyperlinks
       with the URL as literal text (no relationship). Consecutive Hyperlink
       runs are concatenated to handle URLs split across multiple runs.
    """
    import re as _re
    URL_RE = _re.compile(r'https?://\S+')
    seen = set()
    urls = []

    def _add(url):
        url = url.strip()
        if url.startswith('http') and url not in seen:
            seen.add(url)
            urls.append(url)

    # ── 1. Relationship-based <w:hyperlink> elements ──────────────────────
    for hl in el.findall('.//{%s}hyperlink' % NS['w']):
        rid    = hl.get('{%s}id' % NS['r'])
        anchor = hl.get('{%s}anchor' % NS['w'])
        if anchor and anchor.startswith('_Toc'):
            continue
        if rid and rid in rels:
            _add(rels[rid])

    # ── 2. Runs styled with the "Hyperlink" character style ───────────────
    # Iterate paragraph by paragraph so consecutive runs stay together.
    for para in el.findall('.//{%s}p' % NS['w']):
        # Skip paragraphs that are already fully inside a w:hyperlink element
        # (those URLs were captured above)
        if para.getparent() is not None:
            parent_tag = para.getparent().tag.split('}')[-1]
            if parent_tag == 'hyperlink':
                continue

        runs = para.findall('{%s}r' % NS['w'])
        i = 0
        while i < len(runs):
            run = runs[i]
            rStyle = run.find('{%s}rPr/{%s}rStyle' % (NS['w'], NS['w']))
            is_hl_style = (
                rStyle is not None and
                'hyperlink' in (rStyle.get('{%s}val' % NS['w']) or '').lower()
            )
            if is_hl_style:
                # Concatenate all consecutive Hyperlink-styled runs
                parts = []
                while i < len(runs):
                    r2 = runs[i]
                    rs2 = r2.find('{%s}rPr/{%s}rStyle' % (NS['w'], NS['w']))
                    if rs2 is not None and 'hyperlink' in (rs2.get('{%s}val' % NS['w']) or '').lower():
                        t = r2.find('{%s}t' % NS['w'])
                        parts.append((t.text or '') if t is not None else '')
                        i += 1
                    else:
                        break
                full_text = ''.join(parts).strip()
                if URL_RE.match(full_text):
                    _add(full_text)
            else:
                i += 1

    return urls


def _is_section_label_para(para):
    """
    True if a plain (non-heading) paragraph should be used as a section label
    when it immediately precedes a table.  Criteria (all must hold):
      - Has at least one bold run
      - Short enough to be a label (3 to 80 chars) — excludes long body text
      - Does not look like a URL
    """
    content = _txt(para)
    if not content or not (3 < len(content) < 80):
        return False
    if content.startswith('http'):
        return False
    return bool(para.findall('.//{%s}b' % NS['w']))


def extract_from_docx(docx_path, skip_numbering=False):
    """
    Extract hyperlink records from a .docx file.

    skip_numbering=True  — used for LibreOffice-converted .doc files whose
                           auto-numbering may differ from the original document.
    """
    records = []
    with zipfile.ZipFile(docx_path) as z:
        rels    = _load_rels(z)
        with z.open('word/document.xml') as f:
            tree = etree.parse(f)
        body = tree.find('{%s}body' % NS['w'])
        if body is None:
            return records
        num_map = _build_num_map(z, body) if not skip_numbering else {}

    body = tree.find('{%s}body' % NS['w'])
    if body is None:
        return records

    # Section detection — document-agnostic, no hardcoding.
    # Rule 1: Any Heading-styled paragraph updates the current section (always).
    # Rule 2: A bold, short, non-URL paragraph that immediately precedes a table
    #         also updates the section — regardless of whether a heading has been
    #         seen yet. Covers "Program Details", "Document Details" etc.
    children = list(body)

    # Pre-compute which paragraph indices immediately precede a table
    pre_table_idx = set()
    for i, child in enumerate(children):
        if child.tag.split('}')[-1] != 'p' or not _txt(child):
            continue
        for j in range(i + 1, len(children)):
            nt = children[j].tag.split('}')[-1]
            if nt == 'tbl':
                pre_table_idx.add(i)
                break
            if _txt(children[j]):
                break

    current_section = ''
    last_nonempty_para = ''
    heading_seen = False

    for idx, child in enumerate(children):
        tag = child.tag.split('}')[-1]

        if tag == 'p':
            content = _txt(child)
            if content:
                last_nonempty_para = content

            if content and _is_heading_style(child):
                auto_num = num_map.get(child)
                if auto_num:
                    if not re.match(r'^' + re.escape(auto_num) + r'[\s\.]', content):
                        content = auto_num + ' ' + content
                current_section = content
                heading_seen = True

            elif content and idx in pre_table_idx and _is_section_label_para(child):
                current_section = content

        if tag == 'tbl':
            if not heading_seen and last_nonempty_para:
                current_section = last_nonempty_para

            rows = child.findall('{%s}tr' % NS['w'])

            # Detect header row: use row 0 only if it is a true label row.
            # Tables like "Program Details" have no header — row 0 is data.
            header_labels = []
            data_start = 0
            if rows and _is_header_row(rows[0]):
                header_labels = [_txt(tc)
                                 for tc in rows[0].findall('{%s}tc' % NS['w'])]
                data_start = 1   # skip header when assigning row_label
            else:
                # No internal header — look back for a preceding single-row
                # shaded title table (e.g. "Program/Project Name | FA Manage Asset Disposal")
                # whose cell values serve as the column labels for this data table.
                for j in range(idx - 1, -1, -1):
                    prev_tag = children[j].tag.split('}')[-1]
                    prev_txt = _txt(children[j])
                    if prev_tag == 'tbl':
                        prev_rows = children[j].findall('{%s}tr' % NS['w'])
                        if prev_rows and _is_header_row(prev_rows[0]):
                            header_labels = [_txt(tc)
                                             for tc in prev_rows[0].findall('{%s}tc' % NS['w'])]
                        break
                    if prev_txt:   # non-empty paragraph stops lookback
                        break

            for row_idx, row in enumerate(rows):
                cells     = row.findall('{%s}tc' % NS['w'])
                row_label = _txt(cells[0]) if cells else ''

                for col_idx, cell in enumerate(cells):
                    for url in _hl_urls(cell, rels):
                        records.append({
                            'section':       current_section,
                            'table_flag':    'Within Table',
                            'row_label':     row_label,
                            'row_label_col': header_labels[0] if header_labels else '',
                            'url_col':       (header_labels[col_idx]
                                             if col_idx < len(header_labels)
                                             else get_column_letter(col_idx + 1)),
                            'url':           url,
                        })

        elif tag == 'p':
            for url in _hl_urls(child, rels):
                records.append({
                    'section':       current_section,
                    'table_flag':    'Outside Table',
                    'row_label':     '',
                    'row_label_col': '',
                    'url_col':       '',
                    'url':           url,
                })

    return records


# ══════════════════════════════════════════════════════════════════════════
#  Excel writer — single sheet, border-only visual grouping (filterable)
# ══════════════════════════════════════════════════════════════════════════

# ── Style constants matching sample exactly ────────────────────────────────
HDR_FONT  = Font(name='Calibri', size=11, bold=True,  color='FFFFFFFF')
HDR_FILL  = PatternFill('solid', start_color='FF4472C4')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

DAT_FONT  = Font(name='Calibri', size=10, color='FF000000')
DAT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
WHITE_FILL= PatternFill('solid', start_color='FFFFFFFF')
NO_FILL   = PatternFill(fill_type=None)

THIN_BLACK  = Side(style='thin',   color='FF000000')
NO_BORDER   = Side(border_style=None)

def _border(left=False, right=False, top=False, bottom=False):
    return Border(
        left   = THIN_BLACK if left   else NO_BORDER,
        right  = THIN_BLACK if right  else NO_BORDER,
        top    = THIN_BLACK if top    else NO_BORDER,
        bottom = THIN_BLACK if bottom else NO_BORDER,
    )

HDR_BORDER = Border(
    left=THIN_BLACK, right=THIN_BLACK,
    top=THIN_BLACK,  bottom=THIN_BLACK,
)


def _apply_header(ws):
    ws.row_dimensions[1].height = 25
    headers = ['Folder Name', 'Word Document', 'Word Section', 'Table',
               'Row Label', 'Row Label Column', 'URL Column', 'URL',
               'Filter by File', 'Filter by Folder']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border    = HDR_BORDER


def _write_group(ws, excel_row, group, folder, fname):
    """
    Write one file-group. Visual merge via borders:
      A & B cols:
        - ALL rows: right=thin (always visible boundary)
        - FIRST row: top=thin
        - LAST row:  bottom=thin  (sample shows last row also no bottom — 
                     but we add bottom on last to close the block cleanly)
        - MIDDLE rows: no top, no bottom  → seamless block appearance
      C D E H cols: full thin border all sides + white fill (Within Table rows)
      F G cols:     no border, no fill (matches sample exactly)
      Outside Table rows: D/E/F/G/H get no border (only C and H carry full border)
    """
    n = len(group)

    for i, rec in enumerate(group):
        r         = excel_row + i
        is_first  = (i == 0)
        is_last   = (i == n - 1)
        ws.row_dimensions[r].height = 18

        is_table  = rec['table_flag'] == 'Within Table'

        # ── Col A: Folder Name ─────────────────────────────────────────
        # Value only on first row → selecting col A shows Count = distinct folders.
        # Col J (hidden) holds value on every row so the filter still works.
        ca = ws.cell(row=r, column=1, value=folder if is_first else None)
        ca.font      = Font(name='Calibri', size=11)
        ca.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ca.border    = _border(right=True, top=is_first, bottom=is_last)

        # ── Col B: Word Document ───────────────────────────────────────
        # Value only on first row → selecting col B shows Count = distinct files.
        # Col I (hidden) holds value on every row so the filter still works.
        cb = ws.cell(row=r, column=2, value=fname if is_first else None)
        cb.font      = Font(name='Calibri', size=11)
        cb.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cb.border    = _border(right=True, top=is_first, bottom=is_last)

        # ── Col C: Word Section ────────────────────────────────────────
        cc = ws.cell(row=r, column=3, value=rec['section'])
        cc.font      = DAT_FONT
        cc.alignment = DAT_ALIGN
        cc.fill      = WHITE_FILL if is_table else NO_FILL
        cc.border    = _border(left=True, right=True, top=True, bottom=True)

        # ── Col D: Table flag ──────────────────────────────────────────
        cd = ws.cell(row=r, column=4, value=rec['table_flag'])
        cd.font      = DAT_FONT
        cd.alignment = DAT_ALIGN
        cd.fill      = WHITE_FILL if is_table else NO_FILL
        cd.border    = _border(left=True, right=True, top=True, bottom=True) if is_table else _border()

        # ── Col E: Row Label ───────────────────────────────────────────
        ce = ws.cell(row=r, column=5, value=rec['row_label'] or None)
        ce.font      = DAT_FONT
        ce.alignment = DAT_ALIGN
        ce.fill      = WHITE_FILL if is_table else NO_FILL
        ce.border    = _border(left=True, right=True, top=True, bottom=True) if is_table else _border()

        # ── Col F: Row Label Column (no border, no fill — matches sample) ─
        cf = ws.cell(row=r, column=6, value=rec['row_label_col'] or None)
        cf.font      = Font(name='Calibri', size=10, bold=True, color='FF000000')
        cf.alignment = DAT_ALIGN
        cf.fill      = NO_FILL
        cf.border    = _border()

        # ── Col G: URL Column (no border, no fill — matches sample) ───────
        cg = ws.cell(row=r, column=7, value=rec['url_col'] or None)
        cg.font      = Font(name='Calibri', size=10, bold=True, color='FF000000')
        cg.alignment = DAT_ALIGN
        cg.fill      = NO_FILL
        cg.border    = _border()

        # ── Col H: URL ────────────────────────────────────────────────
        ch = ws.cell(row=r, column=8, value=rec['url'])
        ch.font      = DAT_FONT
        ch.alignment = DAT_ALIGN
        ch.fill      = WHITE_FILL if is_table else NO_FILL
        ch.border    = _border(left=True, right=True, top=True, bottom=True) if is_table else _border()

        # ── Col I: "Filter by File" — filename on every row ──────────────
        # Col B (Word Document) has value only on first row → status bar Count
        # shows distinct file count when col B is selected. Users filter by col I
        # to show all rows belonging to a file.
        ci = ws.cell(row=r, column=9, value=fname)
        ci.font      = Font(name='Calibri', size=9, color='FF666666')
        ci.alignment = DAT_ALIGN

        # ── Col J: "Filter by Folder" — folder on every row ───────────────
        cj = ws.cell(row=r, column=10, value=folder)
        cj.font      = Font(name='Calibri', size=9, color='FF666666')
        cj.alignment = DAT_ALIGN

    # ── Row grouping: group non-first rows under the first row ───────────
    # This lets users collapse/expand file groups. Middle rows get outline
    # level 1; the first row (level 0) acts as the group header.
    # summaryBelow=False ensures the group header is the TOP row.
    for i in range(1, n):   # skip first row (i=0)
        ws.row_dimensions[excel_row + i].outline_level = 1
        ws.row_dimensions[excel_row + i].hidden = False

    return excel_row + n


def write_excel(all_records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title        = 'HyperLink_Report'
    ws.freeze_panes = 'A2'

    # Group summary rows are ABOVE their detail rows (standard for this layout)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Autofilter A:J — use col I "Filter by File" to filter all rows of a file
    ws.auto_filter.ref = 'A1:J1'

    # Column widths: A-H visible, I & J hidden (width=0)
    for col, w in zip('ABCDEFGH',
                      [12.54, 45.18, 32.27, 20.0, 21.09, 18.0, 18.0, 70.0]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions['I'].width = 30    # visible filter column for Word Document
    ws.column_dimensions['J'].width = 15    # visible filter column for Folder Name

    _apply_header(ws)

    # Group by (folder, filename) — keep original discovery order
    seen_keys, groups = [], {}
    for rec in all_records:
        key = (rec['folder'], rec['filename'])
        if key not in groups:
            groups[key] = []
            seen_keys.append(key)
        groups[key].append(rec)

    excel_row = 2
    for key in seen_keys:
        folder, fname = key
        excel_row = _write_group(ws, excel_row, groups[key], folder, fname)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'Saved -> {output_path}')


# ══════════════════════════════════════════════════════════════════════════
#  .doc entry point (calls the inline parser above)
# ══════════════════════════════════════════════════════════════════════════

def _extract_from_doc_binary(fpath, folder, fname):
    """Parse a .doc using the inline pure-Python Word Binary Format parser."""
    recs = extract_from_doc(fpath)
    for r in recs:
        r['folder']   = folder
        r['filename'] = fname
    return recs


# ══════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════

def main():
    all_records = []

    files = (list(INPUT_FOLDER.rglob('*.docx')) +
             list(INPUT_FOLDER.rglob('*.doc')))

    if not files:
        print(f'No .docx/.doc files found under {INPUT_FOLDER}')
        return

    for fpath in files:
        try:
            rel    = fpath.relative_to(INPUT_FOLDER)
            folder = rel.parts[0] if len(rel.parts) > 1 else INPUT_FOLDER.name
        except ValueError:
            folder = fpath.parent.name

        fname = fpath.name
        ext   = fpath.suffix.lower()
        print(f'Processing: {fpath.relative_to(INPUT_FOLDER)}')

        try:
            if ext == '.docx':
                recs = extract_from_docx(fpath)
                for r in recs:
                    r['folder']   = folder
                    r['filename'] = fname
                all_records += recs
                if not recs:
                    all_records.append({
                        'folder':        folder,
                        'filename':      fname,
                        'section':       '',
                        'table_flag':    'No Hyperlink found',
                        'row_label':     '',
                        'row_label_col': '',
                        'url_col':       '',
                        'url':           'No Hyperlink found',
                    })
                    print(f'  -> No Hyperlinks found')
                else:
                    t = sum(1 for r in recs if r['table_flag'] == 'Within Table')
                    o = sum(1 for r in recs if r['table_flag'] == 'Outside Table')
                    print(f'  -> {t} table links, {o} outside links')

            elif ext == '.doc':
                # Pure-Python Word Binary Format parser — no installation required
                recs = _extract_from_doc_binary(fpath, folder, fname)
                all_records += recs
                if not recs:
                    all_records.append({
                        'folder': folder, 'filename': fname,
                        'section': '', 'table_flag': 'No Hyperlink found',
                        'row_label': '', 'row_label_col': '',
                        'url_col': '', 'url': 'No Hyperlink found',
                    })
                    print(f'  -> No Hyperlinks found [.doc]')
                else:
                    t = sum(1 for r in recs if r['table_flag'] == 'Within Table')
                    o = sum(1 for r in recs if r['table_flag'] == 'Outside Table')
                    print(f'  -> {t} table links, {o} outside links [.doc]')

        except Exception as exc:
            print(f'  WARNING: {fpath.name}: {exc}')

    print(f'\nTotal records: {len(all_records)}')
    write_excel(all_records, OUTPUT_PATH)


if __name__ == '__main__':
    main()